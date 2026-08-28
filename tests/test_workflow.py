"""High-fidelity integration tests: the whole workflow end-to-end, with a hard
focus on correctness of the spreadsheet - that the RIGHT fields land on the RIGHT
IP row, that per-IP tracking never bleeds across hosts, and that re-scans/updates
preserve everything.

These drive the real parser -> store -> workbook writer -> read-back -> report
paths (no nmap needed) against the bundled sample scan, whose four hosts each
have a distinct fingerprint:

    10.0.10.10  dc01.corp.local  Windows Server 2019  88,389,445,3389  ms17-010
    10.0.10.25  ws01.corp.local  Windows 10 21H2      135,445,3389     (no vulns)
    10.0.20.5   web01            Linux 5.4            22,80,443        4 vulns
    10.0.20.6   web02            Linux 5.4            22,80,21,3306    ftp-anon
"""

import contextlib
import io
import os
import shutil
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from recce import ad
from recce.core import parser
from recce.report.formats import xlsx
from recce.core import tracking as tr
from recce.core.models import Host, Port, Vuln
from recce.report.excel import (build_workbook, read_key_order,
                                 read_workbook_edits, read_workbook_tracking,
                                 update_workbook, STATUS_WIP, STATUS_TODO)
from recce.core.store import Store

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from _workflow_helpers import _LOOT_LINUX, _LOOT_WIN, _GNMAP
from recce.core.targets import _subnet_of

SAMPLE = os.path.join(os.path.dirname(parser.__file__), "sample_scan.xml")

# Ground-truth facts, keyed by IP, for cross-checking the spreadsheet.
FACTS = {
    "10.0.10.10": {"host": "dc01.corp.local", "os": "Windows Server 2019",
                   "ports": [88, 389, 445, 3389], "nvulns": 1},
    "10.0.10.25": {"host": "ws01.corp.local", "os": "Windows 10",
                   "ports": [135, 445, 3389], "nvulns": 0},
    "10.0.20.5": {"host": "web01", "os": "Linux",
                  "ports": [22, 80, 443], "nvulns": 4},
    "10.0.20.6": {"host": "web02", "os": "Linux",
                  "ports": [21, 22, 80, 3306], "nvulns": 1},
}


def sample_hosts():
    hosts = parser.parse_nmap_xml(SAMPLE)
    for h in hosts:
        h.subnet = _subnet_of(h.ip)
        h.enumerated = True
    ad.analyze_hosts(hosts)
    return hosts


def header_index(rows, *must_have):
    """Row index of the real column-header row (the first row that holds every
    token in must_have). A legend/note line can precede the header, so we locate
    it instead of assuming row 0."""
    for i, r in enumerate(rows):
        if all(tok in r for tok in must_have):
            return i
    return 0


def rows_by_ip(sheets, title):
    """Return (header, {ip: [row-as-dict, ...]}) for a sheet with an IP column.

    Skips collapsible group-header band rows (they carry a label in the IP column
    but no Key), so callers only see real data rows keyed by a bare IP."""
    rows = sheets[title]
    hidx = header_index(rows, "IP")
    hdr = rows[hidx]
    ipc = hdr.index("IP")
    kidx = hdr.index("Key") if "Key" in hdr else None
    out: dict = {}
    for r in rows[hidx + 1:]:
        if kidx is not None and (len(r) <= kidx or not r[kidx]):
            continue                       # group-header band row - not data
        if len(r) > ipc and r[ipc]:
            out.setdefault(str(r[ipc]), []).append(dict(zip(hdr, r)))
    return hdr, out
class FullCliRoundTripTest(unittest.TestCase):
    """Drive the real cli report/import functions, as the commands do."""

    def test_report_then_operator_edit_then_import_persists_per_ip(self):
        from recce import cli
        with tempfile.TemporaryDirectory() as d:
            paths = cli._open_paths(d)
            store = Store(paths["db"])
            store.set_meta("engagement", "roundtrip")
            for h in sample_hosts():
                store.upsert_host(h)
            # 1. Generate all reports from the datastore (like `report`).
            cli._generate_reports(store, paths, "roundtrip", quiet=True)
            self.assertTrue(os.path.exists(paths["xlsx"]))
            # 2. Operator edits the workbook: tick Reviewed for web01 only.
            build_workbook(store.all_hosts(), paths["xlsx"],
                           tracking={tr.host_key("10.0.20.5"): (True, "reviewed")},
                           order_map=read_key_order(paths["xlsx"]))
            # 3. Import edits back (like the start of any command).
            cli._import_excel_tracking(store, paths)
            got = store.get_tracking()
            self.assertTrue(got[tr.host_key("10.0.20.5")][0])
            self.assertEqual(got[tr.host_key("10.0.20.5")][1], "reviewed")
            # No other host got reviewed.
            for ip in ("10.0.10.10", "10.0.10.25", "10.0.20.6"):
                self.assertFalse(got.get(tr.host_key(ip), (False, ""))[0])
            # 4. Regenerate -> the tick survives and stays on the right IP.
            cli._generate_reports(store, paths, "roundtrip", quiet=True)
            back = read_workbook_tracking(paths["xlsx"])
            self.assertTrue(back[tr.host_key("10.0.20.5")][0])
            store.close()

    def test_manual_step_override_survives_regeneration(self):
        from recce import cli
        with tempfile.TemporaryDirectory() as d:
            paths = cli._open_paths(d)
            store = Store(paths["db"])
            for h in sample_hosts():
                store.upsert_host(h)
            cli._generate_reports(store, paths, "t", quiet=True)
            # Operator ticks the manual 'Access' box for the DC only.
            key = tr.step_key("access", "10.0.10.10")
            build_workbook(store.all_hosts(), paths["xlsx"],
                           tracking={key: (True, "")},
                           order_map=read_key_order(paths["xlsx"]))
            cli._import_excel_tracking(store, paths)
            self.assertTrue(store.get_tracking()[key][0])
            # Regenerate and confirm it is still checked on the DC row only.
            cli._generate_reports(store, paths, "t", quiet=True)
            sheets = xlsx.read_sheets(paths["xlsx"])
            _hdr, by_ip = rows_by_ip(sheets, "Checklist")
            self.assertEqual(by_ip["10.0.10.10"][0]["Access"], xlsx.CHECK_ON)
            self.assertEqual(by_ip["10.0.20.5"][0]["Access"], xlsx.CHECK_OFF)
            store.close()


class WorkbookStructureTest(unittest.TestCase):
    def test_all_sheets_present_and_key_column_hidden_consistent(self):
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook(sample_hosts(), out)
            sheets = xlsx.read_sheets(out)
        # Always present.
        for name in ("Start Here", "Overview", "Checklist", "Services",
                     "Vulnerabilities", "Services by Product"):
            self.assertIn(name, sheets)
        # Present because the sample has this data (Exploits is skip-if-empty and
        # absent here since searchsploit didn't run).
        for name in ("Databases", "Active Directory", "AD Quick Wins",
                     "Users & Accounts", "Priv-Esc"):
            self.assertIn(name, sheets)
        self.assertNotIn("Exploits", sheets)   # no exploit data -> sheet omitted
        # Tracked sheets carry a Key column so read-back can find every row (the
        # Checklist header sits below its legend line, so locate it, don't assume row 0).
        for name in ("Checklist", "Services", "Vulnerabilities"):
            self.assertIn("Key", sheets[name][header_index(sheets[name], "Key")])

    def test_openpyxl_can_open_the_workbook(self):
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl not installed")
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook(sample_hosts(), out)
            wb = load_workbook(out)
            self.assertIn("Checklist", wb.sheetnames)

    def test_styling_freeze_gridlines_and_severity_contrast(self):
        """The polish pass: identity columns frozen, gridlines off, and critical
        severity is solid with white text (openpyxl reads the applied styles)."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl not installed")
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook(sample_hosts(), out)
            wb = load_workbook(out)
        # Checklist header is on row 2 (a legend line precedes it), so the freeze
        # splits below row 2; 3 identity cols frozen => D3.
        self.assertEqual(wb["Checklist"].freeze_panes, "D3")   # header + 3 id cols
        self.assertEqual(wb["Services"].freeze_panes, "C2")
        self.assertFalse(wb["Checklist"].sheet_view.showGridLines)
        vs = wb["Vulnerabilities"]
        hdr = [c.value for c in vs[1]]
        sev_i = hdr.index("Severity") + 1
        crit = next(vs.cell(row=r, column=sev_i)
                    for r in range(2, vs.max_row + 1)
                    if vs.cell(row=r, column=sev_i).value == "CRITICAL")
        self.assertEqual(crit.fill.fgColor.rgb, "FFC00000")     # solid red
        self.assertEqual(crit.font.color.rgb, "FFFFFFFF")       # white text
        self.assertTrue(crit.font.bold)

    def test_raw_nse_sheet_scopes_host_and_port_scripts_per_host(self):
        from recce.core.models import Host, Port, Script
        hosts = [
            Host(ip="10.0.0.5", hostnames=["a"], ports=[Port(portid=445,
                 service="microsoft-ds", scripts=[Script(id="smb2-security-mode",
                 output="Message signing enabled but not required")])],
                 host_scripts=[Script(id="smb-os-discovery", output="OS: Windows")]),
            Host(ip="10.0.0.6", ports=[Port(portid=21, service="ftp",
                 scripts=[Script(id="ftp-anon", output="Anonymous FTP login allowed")])]),
        ]
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook(hosts, out)
            sheets = xlsx.read_sheets(out)
        self.assertIn("Raw NSE", sheets)
        self.assertIn("Scope", sheets["Raw NSE"][0])          # unified Scope column

        _h, by_ip = rows_by_ip(sheets, "Raw NSE")
        rows5 = {r["Script"]: r for r in by_ip["10.0.0.5"]}
        # Host-level script -> Scope "host"; port script -> Scope "445".
        self.assertEqual(rows5["smb-os-discovery"]["Scope"], "host")
        self.assertEqual(rows5["smb-os-discovery"]["Output"], "OS: Windows")
        self.assertEqual(rows5["smb2-security-mode"]["Scope"], "445")
        # web02's per-port script lands on its own IP, no cross-host bleed.
        rows6 = {r["Script"]: r for r in by_ip["10.0.0.6"]}
        self.assertEqual(rows6["ftp-anon"]["Scope"], "21")
        self.assertNotIn("ftp-anon", rows5)

    def test_design_language_fonts_and_accent(self):
        """Machine data (IP/version) renders monospace with a teal IP accent;
        prose (Product) stays sans - the light HTML-preview design language."""
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl not installed")
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook(sample_hosts(), out)
            wb = load_workbook(out)
        sv = wb["Services"]
        hdr = [c.value for c in sv[1]]
        ki = hdr.index("Key") + 1                          # openpyxl cols are 1-based
        drow = next(r for r in range(2, sv.max_row + 1)
                    if sv.cell(row=r, column=ki).value)   # first real data row

        def cell(name):
            return sv.cell(row=drow, column=hdr.index(name) + 1)
        ip = cell("IP")
        self.assertEqual(ip.font.name, "Consolas")
        self.assertEqual(ip.font.color.rgb, "FF0E6E67")       # teal accent
        self.assertEqual(cell("Version").font.name, "Consolas")  # mono data
        self.assertEqual(cell("Product").font.name, "Calibri")   # prose stays sans
        self.assertEqual(sv.cell(row=1, column=1).fill.fgColor.rgb, "FF0E6E67")  # teal header

    def test_tab_colors_and_overview_nav_links(self):
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl not installed")
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook(sample_hosts(), out)
            wb = load_workbook(out)
        # Tab colours group the tabs into role bands.
        self.assertIsNotNone(wb["Checklist"].sheet_properties.tabColor)
        self.assertIsNotNone(wb["Vulnerabilities"].sheet_properties.tabColor)
        # Overview is a nav hub: a jump bar + clickable totals link to real sheets.
        ov = wb["Overview"]
        targets = {c.hyperlink.location for row in ov.iter_rows()
                   for c in row if c.hyperlink}
        self.assertTrue(any("Checklist" in t for t in targets))
        self.assertTrue(any("Vulnerabilities" in t for t in targets))
        # Every link points at a sheet that actually exists (no dangling jumps).
        present = {f"'{n}'" for n in wb.sheetnames}
        for loc in targets:
            self.assertTrue(loc.split("!")[0] in present, f"dangling link: {loc}")

    def test_credentialed_access_matrix(self):
        from recce.core.models import Host, Port, Vuln

        def cv(t):
            return Vuln(ip="x", port=445, protocol="tcp", script_id="c",
                        title=t, severity="high", source="cred")
        hosts = [
            Host(ip="10.0.10.10", hostnames=["dc01"], os_family="Windows",
                 cred_enumerated=True, ports=[Port(portid=445, service="microsoft-ds")],
                 vulns=[cv("Local admin confirmed - privileged account"),
                        cv("Credential hashes dumped (5 accounts)")]),
            Host(ip="10.0.10.25", hostnames=["ws01"], os_family="Windows",
                 cred_enumerated=True, ports=[Port(portid=445, service="microsoft-ds")],
                 vulns=[cv("Local admin confirmed - user account")]),
        ]
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook(hosts, out)
            rows = xlsx.read_sheets(out)["Overview"]
        # Find the matrix header + the two host rows.
        text = ["|".join(str(c) for c in r) for r in rows]
        self.assertTrue(any("access matrix" in t for t in text))
        dc = next(r for r in rows if r and str(r[0]).startswith("10.0.10.10"))
        ws = next(r for r in rows if r and str(r[0]).startswith("10.0.10.25"))
        # dc01: privileged account is admin + hashes dumped; user account is not admin.
        self.assertEqual([dc[2], dc[3], dc[4]], ["—", "✓", "✓"])
        # ws01: the LOW-PRIV user account is admin (over-privileged) -> flagged.
        self.assertEqual([ws[2], ws[3], ws[4]], ["✓", "—", "—"])

    def test_overview_host_index_deep_links_hit_correct_checklist_rows(self):
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl not installed")
        from recce.report.excel import read_key_order
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook(sample_hosts(), out)

            def _check(path):
                import re as _re
                wb = load_workbook(path)
                ck = wb["Checklist"]
                # The header sits below a legend line - find the row carrying "IP".
                hrow = next(r for r in range(1, ck.max_row + 1)
                            if "IP" in [c.value for c in ck[r]])
                ipc = [c.value for c in ck[hrow]].index("IP") + 1
                # The IP column also carries the collapsible subnet-band labels; keep
                # only the rows whose IP cell is a bare IPv4 (the real host rows).
                ip_row = {}
                for r in range(hrow + 1, ck.max_row + 1):
                    v = ck.cell(row=r, column=ipc).value
                    if isinstance(v, str) and _re.fullmatch(r"\d+\.\d+\.\d+\.\d+", v):
                        ip_row[v] = r
                ov = wb["Overview"]
                deep = {c.value: c.hyperlink.location
                        for row in ov.iter_rows() for c in row
                        if c.hyperlink and c.hyperlink.location.startswith(
                            "'Checklist'!A") and c.value in ip_row}
                # Every host is indexed, and each link targets its real row.
                self.assertEqual(set(deep), set(ip_row))
                for ip, loc in deep.items():
                    self.assertEqual(loc, f"'Checklist'!A{ip_row[ip]}")

            _check(out)
            # Regenerate preserving row order -> links must still be correct.
            build_workbook(sample_hosts(), out, order_map=read_key_order(out))
            _check(out)
            # And after a NEW host is added to an already-seen subnet (it appends at
            # the saved-order tail but the writer re-groups it under its subnet). The
            # linear precompute mis-counted band rows here; the bucketed one must not.
            from recce.report.excel import update_workbook
            from recce.core.models import Host, Port
            extra = Host(ip="10.0.10.99", subnet="10.0.10.0/24", state="up",
                         hostnames=["late01"], os_name="Linux",
                         ports=[Port(portid=22, service="ssh", state="open")])
            update_workbook(out, sample_hosts() + [extra])
            _check(out)

    def test_step_headers_colour_auto_vs_manual(self):
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl not installed")
        from recce.core import tracking as tr
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook(sample_hosts(), out)
            ws = load_workbook(out)["Checklist"]
        # Header is on row 2 (a legend line precedes it) - locate the row with "IP".
        hrow = next(r for r in range(1, ws.max_row + 1)
                    if "IP" in [c.value for c in ws[r]])
        hdr = [c.value for c in ws[hrow]]
        # Access is auto now: recce derives it from credentialed enum (creds/admin/
        # SSH/MSSQL) and it auto-ticks like the other tool phases.
        auto = {"Enumerated", "Vuln-scan", "Web", "DB", "Priv-esc", "Access"}
        manual = {"AD", "Creds", "Lateral"}
        for h in auto:
            c = ws.cell(row=hrow, column=hdr.index(h) + 1)
            self.assertEqual(c.fill.fgColor.rgb, "FF2E7D32", f"{h} should be auto-green")
        for h in manual:
            c = ws.cell(row=hrow, column=hdr.index(h) + 1)
            self.assertEqual(c.fill.fgColor.rgb, "FFC55A11", f"{h} should be manual-amber")
        # Sanity: the split matches the tracking module's source of truth.
        self.assertEqual(manual, {h for h, s in tr.STEP_COLUMNS.items()
                                  if s in tr.MANUAL_STEPS})

    def test_grouped_sheet_has_collapsible_host_bands(self):
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl not installed")
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook(sample_hosts(), out)
            wb = load_workbook(out)
        sv = wb["Services"]
        self.assertEqual(sv.sheet_format.outlineLevelRow, 1)
        self.assertFalse(sv.sheet_properties.outlinePr.summaryBelow)  # header above
        # Detail rows are grouped (outline level 1); host-header rows are level 0.
        levels = {sv.row_dimensions[r].outlineLevel
                  for r in range(2, sv.max_row + 1)}
        self.assertIn(1, levels)                              # some grouped detail
        self.assertIn(0, levels)                              # some header/summary


class ScannerCommandTest(unittest.TestCase):
    """Verify the actual nmap command assembled for each phase (mock _run)."""

    def _capture(self, fn, *a, **k):
        import recce.core.scanner as s
        calls = []
        orig = s._run
        s._run = lambda cmd, timeout=None: (calls.append((cmd, timeout))
                                            or s.RunOutcome(returncode=0))
        try:
            fn(*a, **k)
        finally:
            s._run = orig
        return calls

    def test_full_port_scan_flags(self):
        import recce.core.scanner as s
        with tempfile.TemporaryDirectory() as d:
            calls = self._capture(s.full_port_scan, "1.2.3.4",
                                  os.path.join(d, "p.xml"), s.PROFILES["standard"])
        cmd = calls[0][0]
        self.assertIn("-p-", cmd)                 # full sweep by default
        self.assertIn("--host-timeout", cmd)
        self.assertIn("--max-retries", cmd)
        self.assertIn("1.2.3.4", cmd)
        self.assertIsNotNone(calls[0][1])         # subprocess timeout set

    def test_pn_scan_retries_for_completeness_bounded_by_host_timeout(self):
        # The default sweep must be at least as gentle+reliable as a manual `nmap -p-`:
        # NO --min-rate floor (a floor outpaces a firewall's scan detection -> the source
        # is throttled and ports vanish; PROVEN), and retries at nmap's own -T4 default
        # (6, never the old too-low 3). Dead IPs stay bounded by --host-timeout.
        import copy
        import recce.core.scanner as s
        prof = copy.copy(s.PROFILES["standard"])
        prof.assume_up = True
        with tempfile.TemporaryDirectory() as d:
            calls = self._capture(s.full_port_scan, "1.2.3.4",
                                  os.path.join(d, "p.xml"), prof)
        cmd = calls[0][0]
        self.assertEqual(cmd[cmd.index("--max-retries") + 1], "6")   # >= nmap -T4 default
        self.assertIn("--host-timeout", cmd)                         # bounds dead IPs
        self.assertNotIn("--min-rate", cmd)                          # no overspeed floor

    def test_port_sweep_auto_retries_reliably_on_dropped_probes(self):
        """A rate-limiting network (nmap drops probes) must trigger an automatic
        congestion-adaptive re-scan - no --min-rate floor, more retries, -T3 -
        which is what actually finds the ports (the fast pass under-reports)."""
        import recce.core.scanner as s
        calls = []
        outs = iter([
            s.RunOutcome(returncode=0, stderr="Increasing send delay for 1.2.3.4 "
                         "from 0 to 5 due to 11 out of 11 dropped probes"),
            s.RunOutcome(returncode=0),
        ])
        orig = s._run
        s._run = lambda cmd, timeout=None: (calls.append(cmd) or next(outs))
        try:
            _, issue = s.full_port_scan("1.2.3.4", "/tmp/x.xml", s.ScanProfile())
        finally:
            s._run = orig
        self.assertEqual(len(calls), 2)                       # fast pass + reliable re-scan
        self.assertIn("--min-rate", calls[0])                 # fast pass keeps the floor
        self.assertNotIn("--min-rate", calls[1])              # reliable drops it
        self.assertEqual(calls[1][calls[1].index("--max-retries") + 1], "6")
        self.assertIn("-T3", calls[1])
        # bounded: the adaptive re-scan keeps the SAME --host-timeout as any host
        # (no silent extension), so it can't run for hours - it returns partial.
        self.assertIn("--host-timeout", calls[1])
        self.assertEqual(calls[1][calls[1].index("--host-timeout") + 1],
                         f"{s.ScanProfile().host_timeout}m")
        self.assertTrue(issue and issue.level == "warning")   # rate-limit surfaced

    def test_reliable_flag_drops_min_rate_from_first_pass(self):
        import recce.core.scanner as s
        calls = []
        orig = s._run
        s._run = lambda cmd, timeout=None: (calls.append(cmd)
                                            or s.RunOutcome(returncode=0))
        prof = s.ScanProfile(reliable=True)
        try:
            s.full_port_scan("1.2.3.4", "/tmp/y.xml", prof)
        finally:
            s._run = orig
        self.assertEqual(len(calls), 1)                       # no wasted fast pass
        self.assertNotIn("--min-rate", calls[0])
        self.assertEqual(calls[0][calls[0].index("--max-retries") + 1], "6")

    def test_clean_fast_pass_does_not_rescan(self):
        """No dropped probes -> a single scan, no wasteful reliable re-run."""
        import recce.core.scanner as s
        calls = []
        orig = s._run
        s._run = lambda cmd, timeout=None: (calls.append(cmd)
                                            or s.RunOutcome(returncode=0))
        try:
            s.full_port_scan("1.2.3.4", "/tmp/z.xml", s.ScanProfile())
        finally:
            s._run = orig
        self.assertEqual(len(calls), 1)

    def test_enum_scan_flags(self):
        import recce.core.scanner as s
        with tempfile.TemporaryDirectory() as d:
            calls = self._capture(s.enum_scan, "1.2.3.4", [80, 445],
                                  os.path.join(d, "e.xml"), s.PROFILES["standard"])
        cmd = calls[0][0]
        j = " ".join(cmd)
        self.assertIn("-sV", cmd)
        self.assertIn("--version-intensity", cmd)     # standard = intensity gate
        self.assertIn("--host-timeout", cmd)
        self.assertIn("smb-os-discovery", j)          # AD enrichment scripts
        self.assertIn("80,445", j)                    # exactly the ports given

    def test_udp_liveness_probe_is_a_udp_ping_without_pn(self):
        import recce.core.scanner as s
        orig_root = s._is_root
        s._is_root = lambda: True                     # pretend we have raw-socket caps
        try:
            with tempfile.TemporaryDirectory() as d:
                calls = self._capture(s.udp_liveness_probe, "1.2.3.4",
                                      os.path.join(d, "u.xml"), s.PROFILES["standard"])
        finally:
            s._is_root = orig_root
        cmd = calls[0][0]
        j = " ".join(cmd)
        self.assertIn("-sn", cmd)                     # ping-only (nmap's up verdict)
        self.assertNotIn("-Pn", cmd)                  # NOT -Pn, so up/down is meaningful
        self.assertTrue(any(a.startswith("-PU") for a in cmd))  # UDP ping probes
        self.assertIn("161", j)                       # SNMP among the probed ports
        self.assertIn("53", j)                        # DNS among the probed ports
        self.assertIn("1.2.3.4", cmd)

    def test_udp_liveness_probe_needs_root(self):
        import recce.core.scanner as s
        orig_root = s._is_root
        s._is_root = lambda: False
        try:
            with tempfile.TemporaryDirectory() as d:
                calls = self._capture(s.udp_liveness_probe, "1.2.3.4",
                                      os.path.join(d, "u.xml"), s.PROFILES["standard"])
        finally:
            s._is_root = orig_root
        self.assertEqual(calls, [])                    # no nmap run without root

    def test_vuln_scan_safe_vs_aggressive(self):
        import recce.core.scanner as s
        with tempfile.TemporaryDirectory() as d:
            safe = self._capture(s.vuln_scan, "1.2.3.4", [80],
                                 os.path.join(d, "v.xml"), s.PROFILES["standard"])
            agg = self._capture(s.vuln_scan, "1.2.3.4", [80],
                                os.path.join(d, "v.xml"), s.PROFILES["standard"],
                                aggressive=True)
        safe_j, agg_j = " ".join(safe[0][0]), " ".join(agg[0][0])
        self.assertIn("vuln and safe", safe_j)
        self.assertIn("--version-light", safe[0][0])   # not a full re-scan
        self.assertIn("vuln or vulners", agg_j)
        # KEY FIX: high-value detection scripts that nmap does NOT tag "safe"
        # (ms17-010, heartbleed, vsftpd backdoor) still run in the default scan -
        # no flag needed. --aggressive adds the full intrusive vuln category.
        for script in ("smb-vuln-ms17-010", "ssl-heartbleed", "ftp-vsftpd-backdoor"):
            self.assertIn(script, safe_j)
            self.assertIn(script, agg_j)

    def test_vuln_scan_fast_tier(self):
        import recce.core.scanner as s
        with tempfile.TemporaryDirectory() as d:
            fast = self._capture(s.vuln_scan, "1.2.3.4", [80, 445],
                                 os.path.join(d, "v.xml"), s.PROFILES["standard"],
                                 fast=True)
        fast_j = " ".join(fast[0][0])
        # --fast drops the broad category net + deep enum, keeps top-signal detection.
        self.assertNotIn("vuln and safe", fast_j)
        self.assertNotIn("http-enum", fast_j)          # deep-enum script excluded
        self.assertIn("smb-vuln-ms17-010", fast_j)     # top-signal detection kept
        self.assertIn("ssl-heartbleed", fast_j)
        self.assertIn("90s", fast_j)                   # lighter script-timeout

    def test_enum_deep_scripts_on_standard_not_quick(self):
        import recce.core.scanner as s
        with tempfile.TemporaryDirectory() as d:
            std = self._capture(s.enum_scan, "1.2.3.4", [80],
                                os.path.join(d, "e.xml"), s.PROFILES["standard"])
            quick = self._capture(s.enum_scan, "1.2.3.4", [80],
                                  os.path.join(d, "e2.xml"), s.PROFILES["quick"])
        # Deep service-enum scripts run in enum on standard, dropped on quick.
        self.assertIn("http-enum", " ".join(std[0][0]))
        self.assertNotIn("http-enum", " ".join(quick[0][0]))

    def test_version_all_profile_uses_version_all(self):
        import recce.core.scanner as s
        with tempfile.TemporaryDirectory() as d:
            calls = self._capture(s.enum_scan, "1.2.3.4", [80],
                                  os.path.join(d, "e.xml"), s.PROFILES["thorough"])
        self.assertIn("--version-all", calls[0][0])

    def test_no_ports_writes_empty_xml_and_no_scan(self):
        import recce.core.scanner as s
        with tempfile.TemporaryDirectory() as d:
            xmlp = os.path.join(d, "e.xml")
            calls = self._capture(s.enum_scan, "1.2.3.4", [], xmlp,
                                  s.PROFILES["standard"])
            self.assertEqual(calls, [])                # nothing scanned
            self.assertTrue(os.path.exists(xmlp))      # but a parseable stub exists
            self.assertEqual(parser.parse_nmap_xml(xmlp), [])
def _nmap_xml(ip, ports):
    """Build a minimal, parseable nmap XML for a host with the given ports.

    ports: list of dicts {port, service, product?, version?, scripts?:[(id,out)]}.
    """
    body = [f'<host><status state="up"/>'
            f'<address addr="{ip}" addrtype="ipv4"/><ports>']
    for p in ports:
        body.append(f'<port protocol="tcp" portid="{p["port"]}">'
                    f'<state state="open"/>')
        svc = f'<service name="{p.get("service", "")}"'
        if p.get("product"):
            svc += f' product="{p["product"]}"'
        if p.get("version"):
            svc += f' version="{p["version"]}"'
        body.append(svc + '/>')
        for sid, out in p.get("scripts", []):
            body.append(f'<script id="{sid}" output="{out}"/>')
        body.append('</port>')
    body.append('</ports></host>')
    return '<?xml version="1.0"?><nmaprun start="1">' + "".join(body) + '</nmaprun>'


def _fake_scan(out, ip, ports):
    """Write a canned nmap XML and return the (path, issue) tuple a scan fn does."""
    with open(out, "w") as fh:
        fh.write(_nmap_xml(ip, ports))
    return out, None


class ModelSerializationTest(unittest.TestCase):
    """Host/Domain survive the exact JSON round-trip the store uses (no field loss)."""

    def _rich_host(self):
        from recce.core.models import Account, Exploit, Script
        return Host(
            ip="10.0.0.5", subnet="10.0.0.0/24", hostnames=["h1", "h1.corp"],
            os_name="Linux 5.4", os_family="Linux", os_accuracy=95,
            roles=["Domain Controller"], ntlm={"domain": "CORP"},
            smb_signing="not required", enumerated=True, db_scanned=True,
            privesc_checked=True, cred_enumerated=True, notes="a note",
            ports=[Port(portid=445, service="microsoft-ds", product="Samba",
                        version="4.13", vuln_scanned=True,
                        scripts=[Script(id="smb-os", output="x", elements={"k": "v"})])],
            vulns=[Vuln(ip="10.0.0.5", port=445, protocol="tcp", script_id="v",
                        title="t", severity="high", ids=["CVE-2020-1"],
                        cwes=["CWE-78"], source="version-db", confidence="likely")],
            accounts=[Account(ip="10.0.0.5", source="smb", kind="user", name="a",
                              domain="CORP", rid="500", attrs={"spn": "x"})],
            exploits=[Exploit(ip="10.0.0.5", port=445, edb_id="123",
                              title="e", cves=["CVE-2020-1"])],
            host_scripts=[Script(id="hs", output="o")])

    def test_host_json_roundtrip_via_store_encoding(self):
        import json
        h = self._rich_host()
        h2 = Host.from_json(json.loads(json.dumps(h.to_json())))
        # Scalars + all progress flags.
        for f in ("ip", "subnet", "os_name", "os_family", "smb_signing", "notes",
                  "enumerated", "db_scanned", "privesc_checked", "cred_enumerated"):
            self.assertEqual(getattr(h2, f), getattr(h, f), f)
        self.assertEqual(h2.hostnames, h.hostnames)
        self.assertEqual(h2.roles, h.roles)
        self.assertEqual(h2.ntlm, h.ntlm)
        # Nested structures.
        self.assertEqual(h2.ports[0].scripts[0].elements, {"k": "v"})
        self.assertTrue(h2.ports[0].vuln_scanned)
        self.assertEqual(h2.vulns[0].cwes, ["CWE-78"])       # newest field survives
        self.assertEqual(h2.vulns[0].ids, ["CVE-2020-1"])
        self.assertEqual(h2.accounts[0].attrs, {"spn": "x"})
        self.assertEqual(h2.exploits[0].edb_id, "123")
        self.assertEqual(h2.host_scripts[0].id, "hs")

    def test_domain_json_roundtrip(self):
        from recce.core.models import Domain
        import json
        d = Domain(name="corp.local", netbios="CORP", dc_ips=["10.0.10.10"],
                   anonymous_bind=True, password_policy={"min": 7},
                   trusts=[{"name": "x"}], sources=["nse"])
        d2 = Domain.from_json(json.loads(json.dumps(d.to_json())))
        self.assertEqual(d2.name, "corp.local")
        self.assertEqual(d2.dc_ips, ["10.0.10.10"])
        self.assertTrue(d2.anonymous_bind)
        self.assertEqual(d2.password_policy, {"min": 7})


class PhaseWorkerTest(unittest.TestCase):
    """The real enum/vuln/db/privesc workers, with scanner mocked (no nmap)."""

    def _paths(self, d):
        from recce import cli
        return cli._open_paths(d)

    def test_enum_worker_folds_ports_flags_and_runs_vulndb(self):
        from recce import cli
        import recce.core.scanner as s
        orig = s.enum_scan
        s.enum_scan = lambda ip, ports, out, profile, creds=None: _fake_scan(
            out, ip, [{"port": 21, "service": "ftp", "product": "vsftpd",
                       "version": "2.3.4"}])
        try:
            with tempfile.TemporaryDirectory() as d:
                paths = self._paths(d)
                host, issues = cli._enum_worker(
                    "10.0.0.5", s.PROFILES["standard"], paths, None,
                    {"10.0.0.5": [21]}, {"10.0.0.5": "10.0.0.0/24"})
        finally:
            s.enum_scan = orig
        self.assertTrue(host.enumerated)
        self.assertEqual([p.portid for p in host.open_ports], [21])
        self.assertEqual(host.subnet, "10.0.0.0/24")
        # vulndb ran inside the worker and flagged the vsftpd backdoor.
        self.assertTrue(any("vsftpd 2.3.4 backdoor" in v.title for v in host.vulns))
        self.assertEqual(issues, [])

    def test_vuln_worker_merges_findings_and_marks_ports(self):
        from recce import cli
        import recce.core.scanner as s
        orig = s.vuln_scan
        s.vuln_scan = lambda ip, ports, out, profile, creds=None, aggressive=False, \
            fast=False, skip_enum_scripts=False: \
            _fake_scan(out, ip, [{"port": 80, "service": "http", "scripts": [
                ("http-vuln-x", "VULNERABLE: demo issue State: VULNERABLE")]}])
        try:
            with tempfile.TemporaryDirectory() as d:
                paths = self._paths(d)
                host = Host(ip="10.0.0.5", ports=[Port(portid=80, service="http",
                            state="open")])
                host, issues = cli._vuln_worker(
                    host, [80], s.PROFILES["standard"], paths, None,
                    aggressive=False, use_ss=False, use_probes=False)
        finally:
            s.vuln_scan = orig
        self.assertTrue(host.ports[0].vuln_scanned)          # port marked done
        self.assertTrue(any("http-vuln-x" in (v.script_id or "") for v in host.vulns))
        self.assertEqual(issues, [])

    def test_db_worker_sets_db_scanned(self):
        from recce import cli
        import recce.core.scanner as s
        orig = s.nse_scan
        s.nse_scan = lambda ip, ports, out, profile, scripts, creds=None: _fake_scan(
            out, ip, [{"port": 3306, "service": "mysql"}])
        try:
            with tempfile.TemporaryDirectory() as d:
                paths = self._paths(d)
                host = Host(ip="10.0.0.5", ports=[Port(portid=3306, service="mysql",
                            state="open")])
                host, issues = cli._db_worker(host, [3306], s.PROFILES["standard"],
                                              paths, None, aggressive=False,
                                              use_ss=False)
        finally:
            s.nse_scan = orig
        self.assertTrue(host.db_scanned)
        self.assertTrue(host.ports[0].vuln_scanned)
        self.assertEqual(issues, [])

    def test_privesc_worker_returns_host_and_issues(self):
        from recce import cli
        import recce.core.scanner as s
        orig = s.nse_scan
        s.nse_scan = lambda ip, ports, out, profile, scripts, creds=None: _fake_scan(
            out, ip, [{"port": 445, "service": "microsoft-ds"}])
        try:
            with tempfile.TemporaryDirectory() as d:
                paths = self._paths(d)
                host = Host(ip="10.0.0.9", os_family="Windows",
                            ports=[Port(portid=445, service="microsoft-ds",
                                        state="open")])
                host, issues = cli._privesc_worker(host, s.PROFILES["standard"],
                                                   paths, None, aggressive=False)
        finally:
            s.nse_scan = orig
        self.assertEqual(host.ip, "10.0.0.9")
        self.assertIsInstance(issues, list)


class EnvironmentAndTargetsTest(unittest.TestCase):
    def test_check_environment_requires_nmap_and_warns(self):
        import recce.core.scanner as s
        from recce.core.scanner import ScanProfile, ScannerError
        oh, orr = s._have, s._is_root
        try:
            s._have = lambda t: False        # nothing installed
            s._is_root = lambda: False
            with self.assertRaises(ScannerError):
                s.check_environment(ScanProfile())      # nmap missing -> raise
            s._have = lambda t: t == "nmap"             # only nmap present
            warns = s.check_environment(ScanProfile())
            self.assertTrue(any("root" in w.lower() for w in warns))
            # masscan requested but absent -> warn + fall back to nmap.
            prof = ScanProfile(scanner="masscan")
            warns = s.check_environment(prof)
            self.assertEqual(prof.scanner, "nmap")
            self.assertTrue(any("masscan" in w.lower() for w in warns))
        finally:
            s._have, s._is_root = oh, orr

    def test_load_targets_from_file_with_comments_and_cidr(self):
        from recce.core.targets import load_targets
        with tempfile.TemporaryDirectory() as d:
            f = os.path.join(d, "scope.txt")
            with open(f, "w") as fh:
                fh.write("10.0.0.1\n# a comment\n10.0.0.2   # trailing\n"
                         "10.0.1.0/30\n\n")
            hosts, sm, _ = load_targets(["@" + f])
        self.assertIn("10.0.0.1", hosts)
        self.assertIn("10.0.0.2", hosts)
        self.assertIn("10.0.1.1", hosts)                 # expanded from the CIDR
        self.assertEqual(sm["10.0.1.1"], "10.0.1.0/30")  # CIDR line -> subnet label

    def test_missing_target_file_raises(self):
        from recce.core.targets import load_targets
        with self.assertRaises(FileNotFoundError):
            load_targets(["@/no/such/file.txt"])


class TargetingFormE2ETest(unittest.TestCase):
    """Drive the REAL `vulns` command end-to-end (parser -> selection -> workers ->
    store) and prove each targeting form selects EXACTLY the right stored hosts.

    This locks in the core promise that every phase works on a single IP, several
    IPs, a dash range, a whole subnet, an @file, or 'everything' - using the same
    CLI grammar, with no cross-host bleed. The scanner is mocked so no nmap runs;
    the selection layer under test is entirely real.

    Seeded scope (from the bundled sample): 10.0.10.10, 10.0.10.25 (10.0.10.0/24)
    and 10.0.20.5, 10.0.20.6 (10.0.20.0/24).
    """

    ALL = {"10.0.10.10", "10.0.10.25", "10.0.20.5", "10.0.20.6"}

    def _run_vulns(self, targets):
        """Run `vulns <targets>` against a freshly seeded store; return the set of
        IPs that actually got vuln-scanned (i.e. the hosts the phase selected)."""
        from recce import cli
        import recce.core.scanner as s

        d = tempfile.mkdtemp()
        self.addCleanup(lambda: shutil.rmtree(d, ignore_errors=True))
        paths = cli._open_paths(d)
        seed = Store(paths["db"])
        seed.set_meta("engagement", "e2e")
        for h in sample_hosts():
            seed.upsert_host(h)
        seed.close()

        def fake_vuln(ip, portids, out, profile, creds=None, aggressive=False,
                      fast=False, skip_enum_scripts=False):
            # Echo the requested ports back as an open-port XML the worker parses.
            return _fake_scan(out, ip, [{"port": p, "service": "tcp"}
                                        for p in portids])

        oc, ov, ou = s.check_environment, s.vuln_scan, s.udp_scan
        s.check_environment = lambda profile: []          # no nmap/root needed
        s.vuln_scan = fake_vuln
        s.udp_scan = lambda ip, out, profile: _fake_scan(out, ip, [])
        argv = ["vulns", *targets, "-o", d,
                "--no-searchsploit", "--no-probes", "--workers", "1"]
        try:
            with contextlib.redirect_stdout(io.StringIO()):
                rc = cli.main(argv)
        finally:
            s.check_environment, s.vuln_scan, s.udp_scan = oc, ov, ou
        self.assertEqual(rc, 0)

        store = Store(paths["db"])
        try:
            scanned = {h.ip for h in store.all_hosts()
                       if any(p.vuln_scanned for p in h.open_ports)}
            # Every seeded host must still be present (selection never drops rows).
            self.assertEqual({h.ip for h in store.all_hosts()}, self.ALL)
        finally:
            store.close()
        return scanned

    def test_single_ip(self):
        self.assertEqual(self._run_vulns(["10.0.10.10"]), {"10.0.10.10"})

    def test_several_ips(self):
        self.assertEqual(self._run_vulns(["10.0.10.10", "10.0.20.6"]),
                         {"10.0.10.10", "10.0.20.6"})

    def test_dash_range(self):
        # .10-.25 covers both 10.0.10.x hosts and nothing in 10.0.20.x.
        self.assertEqual(self._run_vulns(["10.0.10.10-25"]),
                         {"10.0.10.10", "10.0.10.25"})

    def test_range_excludes_outside(self):
        # A range that stops before .25 must NOT pick it up.
        self.assertEqual(self._run_vulns(["10.0.10.1-15"]), {"10.0.10.10"})

    def test_whole_subnet(self):
        self.assertEqual(self._run_vulns(["10.0.20.0/24"]),
                         {"10.0.20.5", "10.0.20.6"})

    def test_mixed_single_and_subnet(self):
        self.assertEqual(self._run_vulns(["10.0.10.25", "10.0.20.0/24"]),
                         {"10.0.10.25", "10.0.20.5", "10.0.20.6"})

    def test_at_file(self):
        d = tempfile.mkdtemp()
        self.addCleanup(lambda: shutil.rmtree(d, ignore_errors=True))
        f = os.path.join(d, "scope.txt")
        with open(f, "w") as fh:
            fh.write("10.0.10.10\n"
                     "10.0.20.0/24   # the linux subnet\n")
        self.assertEqual(self._run_vulns(["@" + f]),
                         {"10.0.10.10", "10.0.20.5", "10.0.20.6"})

    def test_empty_targets_selects_everything(self):
        self.assertEqual(self._run_vulns([]), self.ALL)


class UsabilityAndDiscoveryTest(unittest.TestCase):
    def test_pn_alias_parses(self):
        from recce import cli
        for argv in (["enum", "10.0.0.0/24", "-Pn"],
                     ["enum", "10.0.0.0/24", "--no-discovery"],
                     ["scan", "10.0.0.5", "-Pn"]):
            self.assertTrue(cli.build_arg_parser().parse_args(argv).no_discovery, argv)

    def test_bare_recce_prints_quickstart_not_error(self):
        from recce import cli
        buf = io.StringIO()
        with contextlib.redirect_stdout(buf):
            rc = cli.main([])
        self.assertEqual(rc, 0)                        # no argparse error exit
        out = buf.getvalue()
        self.assertIn("recce enum", out)
        self.assertIn("-Pn", out)                      # the ping-blocking hint

    def test_zero_discovery_auto_falls_back_to_pn(self):
        # The killer field bug: hosts that block ping got dropped -> zero ports.
        # Now 0 discovery responses must auto-fall-back to scanning all as up.
        from recce import cli
        import recce.core.scanner as s
        from recce.core.store import Store

        def empty_disc(tf, out):
            with open(out, "w") as fh:
                fh.write('<?xml version="1.0"?><nmaprun></nmaprun>')
            return out, None

        def fps(ip, out, profile):
            with open(out, "w") as fh:
                fh.write(f'<?xml version="1.0"?><nmaprun><host><status state="up"/>'
                         f'<address addr="{ip}" addrtype="ipv4"/><ports><port '
                         f'protocol="tcp" portid="445"><state state="open"/>'
                         f'<service name="microsoft-ds"/></port></ports></host></nmaprun>')
            return out, None

        def enum(ip, ports, out, profile, creds=None):
            return fps(ip, out, profile)

        def empty_udp(ip, out, profile):
            with open(out, "w") as fh:
                fh.write('<?xml version="1.0"?><nmaprun></nmaprun>')
            return out, None

        saved = (s.check_environment, s.discover_hosts, s.full_port_scan, s.enum_scan,
                 s.udp_basic_scan, s.verify_port_scan)
        s.check_environment = lambda p: []
        s.discover_hosts, s.full_port_scan, s.enum_scan = empty_disc, fps, enum
        s.udp_basic_scan = empty_udp
        # The zero-discovery fallback forces verify_all=True, so every live host gets an
        # independent congestion-adaptive re-sweep. Stub it (same as full_port_scan) or
        # the test shells out to a real `nmap -p- --host-timeout 20m` against dead lab IPs.
        s.verify_port_scan = fps
        try:
            with tempfile.TemporaryDirectory() as d:
                buf = io.StringIO()
                with contextlib.redirect_stdout(buf):
                    rc = cli.main(["enum", "10.0.10.10", "10.0.10.11",
                                   "-o", d, "--workers", "1"])
                self.assertEqual(rc, 0)
                self.assertIn("Falling back to -Pn", buf.getvalue())
                hosts = Store(os.path.join(d, "results.sqlite")).all_hosts()
                # Both ping-blocking hosts still got enumerated with their ports.
                self.assertEqual(len(hosts), 2)
                self.assertTrue(all(h.open_ports for h in hosts))
        finally:
            (s.check_environment, s.discover_hosts, s.full_port_scan, s.enum_scan,
             s.udp_basic_scan, s.verify_port_scan) = saved


class PhaseIdempotencyTest(unittest.TestCase):
    """Re-running a phase must NOT duplicate rows. Guards the core store-merge
    contract: hosts/vulns/accounts/exploits/issues dedupe on re-scan."""

    def _counts(self, db):
        s = Store(db)
        try:
            hosts = s.all_hosts()
            return {
                "hosts": len(hosts),
                "vulns": sum(len(h.vulns) for h in hosts),
                "accounts": sum(len(h.accounts) for h in hosts),
                "exploits": sum(len(h.exploits) for h in hosts),
                "issues": s.count_issues().get("total", 0),
            }
        finally:
            s.close()

    def test_rerunning_vulns_does_not_duplicate(self):
        from recce import cli
        import recce.core.scanner as s

        d = tempfile.mkdtemp()
        self.addCleanup(lambda: shutil.rmtree(d, ignore_errors=True))
        paths = cli._open_paths(d)
        seed = Store(paths["db"])
        seed.set_meta("engagement", "idem")
        for h in sample_hosts():
            seed.upsert_host(h)
        seed.close()

        def fake_vuln(ip, portids, out, profile, creds=None, aggressive=False,
                      fast=False, skip_enum_scripts=False):
            # Emit a real NSE finding + an issue every run, so a broken dedup WOULD
            # grow the counts.
            return _fake_scan(out, ip, [{"port": 80, "service": "http", "scripts": [
                ("http-vuln-x", "VULNERABLE: demo State: VULNERABLE")]}])

        oc, ov, ou = s.check_environment, s.vuln_scan, s.udp_scan
        s.check_environment = lambda profile: []
        s.vuln_scan = fake_vuln
        s.udp_scan = lambda ip, out, profile: _fake_scan(out, ip, [])
        argv = ["vulns", "-o", d, "--no-searchsploit", "--no-probes", "--workers", "1"]
        try:
            with contextlib.redirect_stdout(io.StringIO()):
                self.assertEqual(cli.main(argv), 0)
            first = self._counts(paths["db"])
            with contextlib.redirect_stdout(io.StringIO()):
                self.assertEqual(cli.main(argv), 0)          # SAME phase again
            second = self._counts(paths["db"])
        finally:
            s.check_environment, s.vuln_scan, s.udp_scan = oc, ov, ou
        self.assertEqual(first, second, f"re-run changed counts: {first} -> {second}")
        self.assertGreater(first["vulns"], 0)                # actually produced findings

    def test_store_merge_is_idempotent_for_all_collections(self):
        # Upserting the identical rich host twice must not grow any collection.
        from recce.core.models import Account, Exploit, Script
        d = tempfile.mkdtemp()
        self.addCleanup(lambda: shutil.rmtree(d, ignore_errors=True))
        s = Store(os.path.join(d, "r.sqlite"))
        h = Host(
            ip="10.0.0.5", os_family="Linux",
            ports=[Port(portid=445, state="open", service="smb",
                        scripts=[Script(id="smb-os", output="o")])],
            vulns=[Vuln(ip="10.0.0.5", port=445, protocol="tcp", script_id="v",
                        title="t", severity="high")],
            accounts=[Account(ip="10.0.0.5", source="nxc", kind="user", name="a",
                              domain="C", rid="500")],
            exploits=[Exploit(ip="10.0.0.5", port=445, edb_id="1", title="e")],
            host_scripts=[Script(id="hs", output="o")],
            local_findings=[{"category": "sudo", "vector": "NOPASSWD"}])
        s.upsert_host(h)
        s.upsert_host(Host.from_json(h.to_json()))          # identical, again
        got = s.get_host("10.0.0.5")
        s.close()
        self.assertEqual(len(got.vulns), 1)
        self.assertEqual(len(got.accounts), 1)
        self.assertEqual(len(got.exploits), 1)
        self.assertEqual(len(got.host_scripts), 1)
        self.assertEqual(len(got.local_findings), 1)
        self.assertEqual(len(got.ports), 1)
class ProgressAndAuthTest(unittest.TestCase):
    def test_fmt_dur(self):
        from recce import cli
        self.assertEqual(cli._fmt_dur(45), "45s")
        self.assertEqual(cli._fmt_dur(200), "3m20s")
        self.assertEqual(cli._fmt_dur(3660), "1h01m")

    def test_progress_has_pct_and_eta(self):
        from recce import cli
        import time
        s = cli._progress(2, 10, time.monotonic() - 4)
        self.assertIn("20%", s)
        self.assertIn("ETA", s)

    def test_auth_cell(self):
        from recce import cli
        self.assertEqual(cli._auth_cell(None), "-")
        self.assertEqual(cli._auth_cell({"tried": True, "auth": False}), "FAIL")
        self.assertEqual(cli._auth_cell({"tried": True, "auth": True}), "OK")
        self.assertEqual(cli._auth_cell({"tried": True, "auth": True, "admin": True}),
                         "OK (admin)")
        # A tool/connection error is ERR, not FAIL (not a credential problem).
        self.assertEqual(cli._auth_cell({"tried": True, "auth": False, "error": True}),
                         "ERR")

    def test_auth_table_prints_rows_and_flags_fail(self):
        from recce import cli
        rows = [("10.0.0.5", {"user": {"tried": True, "auth": True, "admin": True}}),
                ("10.0.0.9", {"user": {"tried": True, "auth": False}})]
        buf = io.StringIO()
        with contextlib.redirect_stdout(buf):
            cli._print_auth_table(rows)
        out = buf.getvalue()
        self.assertIn("USER ACCT", out)
        self.assertIn("OK (admin)", out)
        self.assertIn("FAIL", out)

    def test_no_yes_flag_and_ingest_present(self):
        from recce import cli
        p = cli.build_arg_parser()
        # ingest is a registered command...
        self.assertEqual(p.parse_args(["ingest", "x.txt"]).command, "ingest")
        # ...and the authorization --yes flag is gone.
        with self.assertRaises(SystemExit):
            p.parse_args(["enum", "1.2.3.4", "--yes"])


class StoreFixesTest(unittest.TestCase):
    def test_corrupt_db_raises_storeerror(self):
        from recce.core.store import Store, StoreError
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "results.sqlite")
            with open(p, "w") as fh:
                fh.write("not a sqlite database, just garbage")
            with self.assertRaises(StoreError):
                Store(p)

    def test_corrupt_db_gives_clean_message_not_traceback(self):
        # A carried-over corrupt DB on `report`/`status` must exit 1 with a clean
        # message, never a raw traceback (the "first command after transfer" case).
        from recce import cli
        with tempfile.TemporaryDirectory() as d:
            os.makedirs(os.path.join(d, "raw"), exist_ok=True)
            with open(os.path.join(d, "results.sqlite"), "w") as fh:
                fh.write("garbage, not a database")
            for command in ("report", "status"):
                buf = io.StringIO()
                with contextlib.redirect_stdout(buf):
                    rc = cli.main([command, "-o", d])
                self.assertEqual(rc, 1, command)
                out = buf.getvalue()
                self.assertIn("corrupt or unreadable", out)
                self.assertNotIn("Traceback", out)

    def test_distance_preserved_through_merge(self):
        from recce.core.store import Store
        with tempfile.TemporaryDirectory() as d:
            s = Store(os.path.join(d, "r.sqlite"))
            s.upsert_host(Host(ip="10.0.0.5", distance=3))
            # A second scan of the same host without distance must not zero it.
            s.upsert_host(Host(ip="10.0.0.5", os_name="Linux", os_accuracy=95))
            self.assertEqual(s.get_host("10.0.0.5").distance, 3)

    def test_rerun_does_not_duplicate_issues(self):
        from recce import cli
        from recce.core.store import Store
        with tempfile.TemporaryDirectory() as d:
            paths = cli._open_paths(d)
            s = Store(paths["db"])
            iss = [{"phase": "vuln-scan", "level": "error", "message": "nmap failed"}]
            cli._record_issues(s, paths, "10.0.0.5", iss)
            cli._record_issues(s, paths, "10.0.0.5", iss)   # re-run, same phase
            self.assertEqual(s.count_issues().get("total"), 1)  # replaced, not doubled
            s.close()


class RunbookSheetTest(unittest.TestCase):
    def test_runbook_sheet_present(self):
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "wb.xlsx")
            build_workbook([], p, meta={"subtitle": "x"})
            try:
                import openpyxl
            except ImportError:
                self.skipTest("openpyxl not installed (test-only dependency)")
            wb = openpyxl.load_workbook(p)
            self.assertIn("Runbook", wb.sheetnames)
            vals = [c for row in wb["Runbook"].iter_rows(values_only=True)
                    for c in row if c]
            joined = " ".join(str(v) for v in vals)
            self.assertIn("enum", joined)
            self.assertIn("ingest", joined)
            self.assertIn("--fast", joined)


class EntryPointTest(unittest.TestCase):
    def test_module_entrypoint_runs(self):
        import subprocess
        root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        env = dict(os.environ, PYTHONPATH=root)
        r = subprocess.run([sys.executable, "-m", "recce", "--version"],
                           capture_output=True, text=True, env=env, timeout=30)
        self.assertEqual(r.returncode, 0)
        self.assertIn("recce", (r.stdout + r.stderr).lower())


class EnumRobustnessTest(unittest.TestCase):
    """The enum phase must be robust host-by-host: one host that crashes, times
    out, or returns hostile data can never abort the run or corrupt the workbook.
    """

    def _args(self, d):
        from types import SimpleNamespace
        a = SimpleNamespace(workers=4, refresh_every=0, title="T", resume=False,
                            user=None, hash=None, domain=None, output_dir=d)
        setattr(a, "pass", None)
        for k in ("ssh_user", "ssh_pass", "ssh_key", "admin_user", "admin_pass",
                  "admin_domain", "dc_ip"):
            setattr(a, k, None)
        return a

    def test_one_bad_host_does_not_abort_run_or_corrupt_workbook(self):
        from recce import cli
        from recce.core import scanner
        from recce.report.formats import xlsx
        import zipfile
        d = tempfile.mkdtemp()
        self.addCleanup(shutil.rmtree, d, ignore_errors=True)
        paths = cli._open_paths(d)
        store = cli._open_store(paths["db"])
        store.set_scope("10.0.0.0/24", 254)

        def fake_worker(ip, profile, paths, creds, port_map, subnet_map,
                        active_probe=True, disc_reason="", provided_name=""):
            if ip == "10.0.0.11":            # worker raises
                raise RuntimeError("boom")
            if ip == "10.0.0.12":            # timed out -> None + issue
                return None, [{"phase": "enum", "level": "error",
                               "message": "host timeout"}]
            if ip == "10.0.0.13":            # hostile data: control chars, many ports
                return Host(ip=ip, subnet="10.0.0.0/24", enumerated=True,
                            hostnames=["odd\x01\x1f"], ports=[
                                Port(portid=n, service="x\x02y", state="open")
                                for n in range(1, 60)]), []
            return Host(ip=ip, subnet="10.0.0.0/24", enumerated=True,
                        ports=[Port(portid=445, service="microsoft-ds",
                                    state="open")]), []

        orig = cli._enum_worker
        cli._enum_worker = fake_worker
        try:
            live = ["10.0.0.10", "10.0.0.11", "10.0.0.12", "10.0.0.13", "10.0.0.14"]
            with contextlib.redirect_stdout(io.StringIO()):
                cli._phase_enum(store, paths, self._args(d),
                                scanner.PROFILES["standard"],
                                {"10.0.0.10": "10.0.0.0/24"}, live,
                                {i: [] for i in live})
                cli._generate_reports(store, paths, "T", quiet=True)
        finally:
            cli._enum_worker = orig

        ips = {h.ip for h in store.all_hosts()}
        # good + hostile-but-valid hosts persist; crashed/timed-out do not
        self.assertEqual(ips, {"10.0.0.10", "10.0.0.13", "10.0.0.14"})
        issue_ips = {i["ip"] for i in store.get_issues()}
        self.assertTrue({"10.0.0.11", "10.0.0.12"} <= issue_ips)
        # the workbook is valid despite the failures + control chars
        self.assertTrue(zipfile.is_zipfile(paths["xlsx"]))
        self.assertIn("Checklist", xlsx.read_sheets(paths["xlsx"]))
        store.close()

    def test_persist_failure_on_one_host_does_not_abort_the_phase(self):
        from recce import cli
        from recce.core import scanner
        d = tempfile.mkdtemp()
        self.addCleanup(shutil.rmtree, d, ignore_errors=True)
        paths = cli._open_paths(d)
        store = cli._open_store(paths["db"])
        store.set_scope("10.0.0.0/24", 254)

        def good_worker(ip, *a, **k):
            return (Host(ip=ip, subnet="10.0.0.0/24", enumerated=True,
                         ports=[Port(portid=445, service="microsoft-ds",
                                     state="open")]), [])
        # The datastore rejects exactly one host's write (a lock outlasting the
        # busy_timeout, say); every other host must still persist.
        real_upsert = store.upsert_host

        def flaky_upsert(host):
            if host.ip == "10.0.0.12":
                raise RuntimeError("database is locked")
            return real_upsert(host)
        store.upsert_host = flaky_upsert

        orig = cli._enum_worker
        cli._enum_worker = good_worker
        try:
            live = ["10.0.0.10", "10.0.0.11", "10.0.0.12", "10.0.0.13"]
            with contextlib.redirect_stdout(io.StringIO()):
                cli._phase_enum(store, paths, self._args(d),
                                scanner.PROFILES["standard"],
                                {"10.0.0.10": "10.0.0.0/24"}, live,
                                {i: [] for i in live})
        finally:
            cli._enum_worker = orig
        store.upsert_host = real_upsert
        ips = {h.ip for h in store.all_hosts()}
        self.assertEqual(ips, {"10.0.0.10", "10.0.0.11", "10.0.0.13"})
        self.assertIn("10.0.0.12", {i["ip"] for i in store.get_issues()})
        store.close()

    def test_zero_port_host_gets_verification_rescan(self):
        """A host the fast pass found 0 ports on is re-verified with an adaptive
        re-scan (discovered-live always; -Pn only with --verify-all), so a missed
        sweep isn't silently trusted as 'no ports'."""
        from recce import cli
        from recce.core import scanner
        from recce.core.models import Host
        calls = {"verify": 0}
        saved = (cli._ports_for_host, cli._fold_host, scanner.full_port_scan,
                 scanner.verify_port_scan, scanner.enum_scan, cli.np.parse_nmap_xml)

        def fake_ports(path, ip):
            return [80, 443] if "verify" in path else []   # fast=0, verify finds some

        def fake_verify(ip, out, profile):
            calls["verify"] += 1
            return out, None
        cli._ports_for_host = fake_ports
        cli._fold_host = lambda ip, parsed, sm: Host(ip=ip, subnet="s")
        scanner.full_port_scan = lambda ip, out, profile: (out, None)
        scanner.verify_port_scan = fake_verify
        scanner.enum_scan = lambda ip, ports, out, profile, creds=None: (out, None)
        cli.np.parse_nmap_xml = lambda p: []
        try:
            d = tempfile.mkdtemp()
            self.addCleanup(shutil.rmtree, d, ignore_errors=True)
            paths = {"raw": d}
            # discovered-live 0-port host -> verified
            cli._enum_worker("1.2.3.4", scanner.ScanProfile(ping_discovery=True,
                             udp_basic=False), paths, None, None, {})
            self.assertEqual(calls["verify"], 1)
            # -Pn (assume-up) without --verify-all -> NOT re-scanning every dead IP
            calls["verify"] = 0
            cli._enum_worker("1.2.3.5",
                             scanner.ScanProfile(ping_discovery=False, verify_all=False, udp_basic=False),
                             paths, None, None, {})
            self.assertEqual(calls["verify"], 0)
            # -Pn WITH --verify-all -> verified
            cli._enum_worker("1.2.3.6",
                             scanner.ScanProfile(ping_discovery=False, verify_all=True, udp_basic=False),
                             paths, None, None, {})
            self.assertEqual(calls["verify"], 1)
            # verify disabled -> never
            calls["verify"] = 0
            cli._enum_worker("1.2.3.7",
                             scanner.ScanProfile(ping_discovery=True, verify=False, udp_basic=False),
                             paths, None, None, {})
            self.assertEqual(calls["verify"], 0)
        finally:
            (cli._ports_for_host, cli._fold_host, scanner.full_port_scan,
             scanner.verify_port_scan, scanner.enum_scan,
             cli.np.parse_nmap_xml) = saved

    def test_truncated_sweep_incomplete_flag_round_trips_and_clears_on_recompletion(self):
        """A truncated sweep flags the host incomplete_scan (persisted); a later
        complete sweep clears it, and ports union across the two scans."""
        from recce.core.store import Store
        from recce.core.models import Host, Port
        d = tempfile.mkdtemp()
        self.addCleanup(shutil.rmtree, d, ignore_errors=True)
        st = Store(os.path.join(d, "s.sqlite"))
        st.upsert_host(Host(ip="1.2.3.4", incomplete_scan=True,
                            ports=[Port(portid=80, service="http", state="open")]))
        self.assertTrue(st.all_hosts()[0].incomplete_scan)   # persisted
        # a later, complete sweep of the same host
        st.upsert_host(Host(ip="1.2.3.4", incomplete_scan=False,
                            ports=[Port(portid=443, service="https", state="open")]))
        h = st.all_hosts()[0]
        self.assertFalse(h.incomplete_scan)                  # complete once either finished
        self.assertEqual({p.portid for p in h.open_ports}, {80, 443})   # union
        st.close()

    def test_corrupt_existing_workbook_is_regenerated_not_fatal(self):
        from recce.report.excel import update_workbook
        from recce.report.formats import xlsx
        import zipfile
        d = tempfile.mkdtemp()
        self.addCleanup(shutil.rmtree, d, ignore_errors=True)
        p = os.path.join(d, "enumeration.xlsx")
        with open(p, "wb") as f:            # a truncated / non-xlsx file
            f.write(b"PK\x03\x04 not a real workbook \xff\x00")
        hosts = [Host(ip="10.0.0.5", subnet="10.0.0.0/24", enumerated=True,
                      ports=[Port(portid=80, service="http", state="open")])]
        update_workbook(p, hosts, meta={"subtitle": "t"}, tracking={})
        self.assertTrue(zipfile.is_zipfile(p))
        self.assertIn("Checklist", xlsx.read_sheets(p))


if __name__ == "__main__":
    unittest.main()
