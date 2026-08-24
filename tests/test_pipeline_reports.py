"""Offline tests split out of tests/test_pipeline.py.

Every test class here is what the original monolith called it. Shared
helpers (header_index, _docx_text, _self_response) live in _pipeline_helpers."""
"""Offline tests for the enumeration pipeline (no network / nmap needed)."""

import contextlib
import io
import os
import stat
import sys
import tempfile
import unittest
from types import SimpleNamespace
from unittest import mock

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from recce import ad, exploits, parser, scanner
from recce import tracking as tr
from recce import xlsx
from recce.models import Account, Host, Port, Script, Vuln
from recce.report_excel import (build_workbook, read_workbook_tracking,
                                       update_workbook)
from recce.store import Store
from recce.targets import apply_exclusions, load_targets

SAMPLE = os.path.join(os.path.dirname(parser.__file__), "sample_scan.xml")


from _pipeline_helpers import header_index, _docx_text, _self_response, SAMPLE  # noqa: F401





class CoverageTest(unittest.TestCase):
    def setUp(self):
        from recce.targets import _subnet_of
        self.hosts = parser.parse_nmap_xml(SAMPLE)
        for h in self.hosts:
            h.subnet = _subnet_of(h.ip)
        ad.analyze_hosts(self.hosts)

    def test_item_keys_categories(self):
        keys = tr.item_keys(self.hosts)
        self.assertEqual(len(keys["hosts"]), 4)
        self.assertEqual(len(keys["services"]), 14)
        self.assertTrue(keys["quick_wins"])  # DC + relay + smbv1

    def test_coverage_counts(self):
        tracking = {tr.host_key("10.0.10.10"): (True, "done")}
        cov = tr.compute_coverage(self.hosts, tracking)
        self.assertEqual(cov["hosts"]["done"], 1)
        self.assertEqual(cov["hosts"]["total"], 4)
        self.assertEqual(cov["overall"]["done"], 1)

    def test_subnet_coverage(self):
        tracking = {tr.host_key("10.0.10.10"): (True, "")}
        sc = tr.subnet_coverage(self.hosts, tracking)
        self.assertEqual(sc["10.0.10.0/24"]["done"], 1)
        self.assertEqual(sc["10.0.10.0/24"]["total"], 2)




class WorkbookFlowTest(unittest.TestCase):
    """The sheet order follows the engagement flow, with the service deep-dive band
    grouped and the AD cluster kept contiguous."""

    def _build(self):
        from recce.models import Domain, Vuln
        from recce import xlsx

        def V(ip, port, sid, title, sev, src):
            return Vuln(ip=ip, port=port, protocol="tcp", script_id=sid, title=title,
                        severity=sev, source=src, state="finding")
        hosts = [
            Host(ip="10.0.0.11", os_family="Windows", roles=["Domain Controller"],
                 enumerated=True,
                 ports=[Port(portid=445, state="open", service="microsoft-ds"),
                        Port(portid=1433, state="open", service="ms-sql-s",
                             product="Microsoft SQL Server")],
                 vulns=[V("10.0.0.11", 445, "smb:x", "SMB signing not required",
                          "medium", "smb"),
                        V("10.0.0.11", 1433, "mssql:x", "MSSQL sysadmin",
                          "critical", "mssql")]),
            Host(ip="10.0.0.22", os_family="Linux", enumerated=True,
                 ports=[Port(portid=2375, state="open", service="docker")],
                 vulns=[V("10.0.0.22", 2375, "docker:x",
                          "Docker Engine API exposed without authentication",
                          "critical", "docker")]),
        ]
        m = {"targets": [{"ip": "x", "port": 1}], "findings": [], "runbooks": []}
        meta = {"subtitle": "Flow", "mssql": m, "smb": m, "docker": m,
                "ad_bloodhound": {"findings": [{"severity": "high",
                    "title": "Kerberoastable", "principal": "svc", "target": "",
                    "detail": "d", "tool": "t", "command": "c", "remediation": "r",
                    "category": "kerberoast", "cwes": ["CWE-262"]}],
                    "paths": [{"start": "a", "target": "DA", "chain": "x", "length": 1,
                               "who": "a", "steps": [], "any_user": False}],
                    "kerberos": [], "domains": [{"name": "corp.local"}], "stats": {}}}
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook(hosts, out, meta=meta,
                           domains=[Domain(name="corp.local", dc_ips=["10.0.0.11"])])
            return list(xlsx.read_sheets(out).keys()), xlsx.read_sheets(out)

    def test_service_band_and_ad_cluster_are_grouped(self):
        tabs, _ = self._build()
        pos = {t: i for i, t in enumerate(tabs)}
        # Service deep-dive band is contiguous and sits right after Databases.
        self.assertLess(pos["Databases"], pos["MSSQL"])
        self.assertEqual(pos["SMB"], pos["MSSQL"] + 1)
        self.assertEqual(pos["Docker"], pos["SMB"] + 1)
        # The whole service band precedes the AD cluster.
        self.assertLess(pos["Docker"], pos["Active Directory"])
        # AD cluster is contiguous: inventory -> quick wins -> findings -> paths.
        self.assertEqual(pos["AD Quick Wins"], pos["Active Directory"] + 1)
        self.assertEqual(pos["AD Findings"], pos["AD Quick Wins"] + 1)
        self.assertEqual(pos["AD Attack Paths"], pos["AD Findings"] + 1)
        # Exploit/chain the foothold BEFORE post-ex priv-esc.
        self.assertLess(pos["Exploitation"], pos["Priv-Esc"])
        self.assertLess(pos["Attack Path"], pos["Priv-Esc"])

    def test_overview_shows_confirmed_metric_and_nav_matches_order(self):
        tabs, sheets = self._build()
        ov = ["|".join(str(c) for c in r) for r in sheets["Overview"]]
        self.assertTrue(any("Confirmed by recce (prove engine)" in t for t in ov))
        # The jump-bar lists the tabs in the same left-to-right order they appear.
        navrow = next((r for r in sheets["Overview"]
                       if any(str(c).strip() == "Checklist" for c in r)), None)
        self.assertIsNotNone(navrow)
        nav = [str(c).strip() for c in navrow if str(c).strip()
               and str(c).strip() != "Jump to"]
        nav_in_tabs = [t for t in nav if t in tabs]
        self.assertEqual(nav_in_tabs, sorted(nav_in_tabs, key=lambda t: tabs.index(t)))




class PortStatusTest(unittest.TestCase):
    """Per-port tri-state work status on the Services sheet."""

    def _host(self):
        return Host(ip="10.0.0.5", subnet="10.0.0.0/24", enumerated=True,
                    ports=[Port(portid=80, service="http", state="open"),
                           Port(portid=443, service="https", state="open")])

    def test_services_sheet_has_status_column_and_dropdown(self):
        from recce.report_excel import (build_workbook, STATUS_VALUES,
                                         STATUS_TODO)
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook([self._host()], out)
            rows = xlsx.read_sheets(out)["Services"]
            hdr = rows[0]
            self.assertIn("Status", hdr)
            self.assertIn("Notes", hdr)
            si = hdr.index("Status")
            ki = hdr.index("Key")
            # Every port DATA row defaults to "Not started" (skip the collapsible
            # host-header band rows, which carry no Key).
            data_rows = [r for r in rows[1:] if len(r) > ki and r[ki]]
            self.assertTrue(data_rows)
            for r in data_rows:
                self.assertEqual(r[si], STATUS_TODO)
            # The dropdown offers all three states (find the sheet whose
            # data-validation lists them, not merely any sheet mentioning them).
            import zipfile
            listing = ",".join(STATUS_VALUES)
            with zipfile.ZipFile(out) as z:
                xmls = [z.read(n).decode() for n in z.namelist()
                        if "worksheets/sheet" in n]
            self.assertTrue(any(f'<formula1>"{listing}"</formula1>' in x
                                for x in xmls),
                            "Services Status dropdown not found")

    def test_status_roundtrip_and_reviewed_mapping(self):
        from recce.report_excel import (build_workbook, read_workbook_edits,
                                         STATUS_WIP, STATUS_DONE)
        with tempfile.TemporaryDirectory() as d:
            store = Store(os.path.join(d, "t.sqlite"))
            store.upsert_host(self._host())
            paths = {"xlsx": os.path.join(d, "wb.xlsx")}
            k80 = tr.svc_key("10.0.0.5", "tcp", 80)
            k443 = tr.svc_key("10.0.0.5", "tcp", 443)
            # Persist an in-progress port and a done port.
            store.bulk_set_status({k80: (STATUS_WIP, False, "poking at it"),
                                   k443: (STATUS_DONE, True, "")})
            # Regenerate from the store, then read the sheet back.
            build_workbook(store.all_hosts(), paths["xlsx"],
                           tracking=store.get_tracking(),
                           statuses=store.get_statuses())
            edits, statuses = read_workbook_edits(paths["xlsx"])
            self.assertEqual(statuses[k80], STATUS_WIP)
            self.assertEqual(statuses[k443], STATUS_DONE)
            # In-progress is not "reviewed"; done is.
            self.assertFalse(edits[k80][0])
            self.assertTrue(edits[k443][0])
            self.assertEqual(edits[k80][1], "poking at it")
            # Coverage counts only the done port.
            cov = tr.compute_coverage(store.all_hosts(), store.get_tracking())
            self.assertEqual(cov["services"]["done"], 1)
            store.close()

    def test_status_column_not_misread_as_checkbox(self):
        # The Status column sits at index 0 (where a checkbox used to be); a
        # "Not started" cell must not be read as reviewed=True.
        from recce.report_excel import build_workbook, read_workbook_edits
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook([self._host()], out)
            edits, _ = read_workbook_edits(out)
            self.assertFalse(edits[tr.svc_key("10.0.0.5", "tcp", 80)][0])

    def test_status_survives_store_migration(self):
        # A datastore created before the status column still gains it.
        with tempfile.TemporaryDirectory() as d:
            path = os.path.join(d, "old.sqlite")
            import sqlite3
            con = sqlite3.connect(path)
            con.executescript(
                "CREATE TABLE tracking (key TEXT PRIMARY KEY, reviewed INTEGER "
                "DEFAULT 0, notes TEXT DEFAULT '', updated TEXT DEFAULT '');")
            con.commit(); con.close()
            store = Store(path)   # __init__ migrates
            store.bulk_set_status({"svc:x": ("◐ In progress", False, "")})
            self.assertEqual(store.get_statuses()["svc:x"], "◐ In progress")
            store.close()




class LdifParseTest(unittest.TestCase):
    def test_parse_entries_and_base64(self):
        import base64 as b64
        enc = b64.b64encode("héllo".encode()).decode()
        ldif = (
            "dn: CN=svc_sql,DC=corp,DC=local\n"
            "sAMAccountName: svc_sql\n"
            "servicePrincipalName: MSSQLSvc/db01.corp.local:1433\n"
            "userAccountControl: 66048\n"
            f"description:: {enc}\n"
            "\n"
            "dn: CN=alice,DC=corp,DC=local\n"
            "sAMAccountName: alice\n"
            "userAccountControl: 4260352\n"
        )
        entries = ad._parse_ldif(ldif)
        self.assertEqual(len(entries), 2)
        self.assertEqual(entries[0]["sAMAccountName"], ["svc_sql"])
        # internal colon in SPN preserved
        self.assertEqual(entries[0]["servicePrincipalName"], ["MSSQLSvc/db01.corp.local:1433"])
        self.assertEqual(entries[0]["description"], ["héllo"])
        # AS-REP flag (0x400000) set on alice
        acc = ad._acc_from_ldif(entries[1], "10.0.0.1", "corp.local", "user")
        self.assertEqual(acc.attrs.get("asrep_roastable"), "yes")




class WeakConfigFindingTest(unittest.TestCase):
    def setUp(self):
        self.hosts = {h.ip: h for h in parser.parse_nmap_xml(SAMPLE)}

    def _find(self, ip, script_id):
        return next((v for v in self.hosts[ip].vulns if v.script_id == script_id), None)

    def test_ftp_anon_medium(self):
        v = self._find("10.0.20.6", "ftp-anon")
        self.assertIsNotNone(v)
        self.assertEqual(v.severity, "medium")
        self.assertTrue(v.cwes)                     # weak-config carries CWEs
        self.assertEqual(v.source, "config")

    def test_weak_tls_medium(self):
        v = self._find("10.0.20.5", "ssl-enum-ciphers")
        self.assertEqual(v.severity, "medium")

    def test_expired_cert_low(self):
        v = self._find("10.0.20.5", "ssl-cert")
        self.assertEqual(v.severity, "low")

    def test_risky_methods_low(self):
        v = self._find("10.0.20.5", "http-methods")
        self.assertEqual(v.severity, "low")

    def test_cve_still_takes_precedence(self):
        # smb-vuln-ms17-010 stays a CVE finding, not reclassified.
        v = self._find("10.0.10.10", "smb-vuln-ms17-010")
        self.assertEqual(v.severity, "critical")




class DocxWriterTest(unittest.TestCase):
    def test_writer_parts_and_text(self):
        from recce.docx import Document
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "t.docx")
            doc = Document()
            doc.title("Hello")
            doc.heading("Section")
            doc.field("Severity", "HIGH")
            doc.placeholder("do this")
            doc.save(out)
            text, parts = _docx_text(out)
        self.assertIn("[Content_Types].xml", parts)
        self.assertIn("word/document.xml", parts)
        self.assertIn("Hello", text)
        self.assertIn("Severity: HIGH", text)
        self.assertIn("[TESTER: do this]", text)

    def test_design_language_styling(self):
        """Teal accent, coloured/mono field values, teal-tinted evidence block."""
        import zipfile
        from recce.docx import Document
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "s.docx")
            doc = Document()
            doc.title("T")
            doc.field("Severity", "CRITICAL", value_color="C00000")
            doc.field("CVE / References", "CVE-2021-41773", mono=True)
            doc.mono_block("raw evidence line")
            doc.save(out)
            with zipfile.ZipFile(out) as z:
                body = z.read("word/document.xml").decode()
                styles = z.read("word/styles.xml").decode()
        self.assertIn('w:color w:val="0E6E67"', styles)      # teal accent in headings
        self.assertIn('w:color w:val="C00000"', body)        # severity value coloured
        self.assertIn('w:ascii="Consolas"', body)            # mono CVE + evidence
        self.assertIn('w:fill="EDF6F4"', body)               # teal-tinted evidence

    def test_toc_field_and_updatefields_toggle(self):
        """A doc with a TOC emits the field + updateFields (Word rebuilds on open);
        a plain doc emits settings.xml WITHOUT updateFields, so it never prompts."""
        import zipfile
        from recce.docx import Document
        with tempfile.TemporaryDirectory() as d:
            # with TOC
            p1 = os.path.join(d, "toc.docx")
            doc = Document(); doc.title("T"); doc.toc(); doc.heading("A"); doc.save(p1)
            with zipfile.ZipFile(p1) as z:
                body = z.read("word/document.xml").decode()
                s1 = z.read("word/settings.xml").decode()
            self.assertIn('w:fldCharType="begin"', body)
            self.assertIn('TOC \\o', body)
            self.assertIn('<w:updateFields w:val="true"/>', s1)
            # without TOC -> settings present but no updateFields (no open-time prompt)
            p2 = os.path.join(d, "plain.docx")
            doc = Document(); doc.title("T"); doc.heading("A"); doc.save(p2)
            with zipfile.ZipFile(p2) as z:
                s2 = z.read("word/settings.xml").decode()
            self.assertNotIn("updateFields", s2)
            _docx_text(p1); _docx_text(p2)   # both parts well-formed

    def test_table_body_cell_colour(self):
        """body_colors tints an individual body cell (severity ramp on counts)."""
        import zipfile
        from recce.docx import Document
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "tc.docx")
            doc = Document()
            doc.table(["Critical", "Low"], [["3", "1"]],
                      body_colors=[["C00000", "2E5AAC"]])
            doc.save(out)
            with zipfile.ZipFile(out) as z:
                body = z.read("word/document.xml").decode()
            self.assertIn('w:color w:val="C00000"', body)
            self.assertIn('w:color w:val="2E5AAC"', body)
            _docx_text(out)

    def test_image_embed(self):
        import struct
        import binascii
        import zlib
        from recce.docx import Document, _png_size
        sig = b"\x89PNG\r\n\x1a\n"

        def chunk(t, dat):
            return (struct.pack(">I", len(dat)) + t + dat
                    + struct.pack(">I", binascii.crc32(t + dat) & 0xffffffff))
        png = (sig + chunk(b"IHDR", struct.pack(">IIBBBBB", 640, 480, 8, 2, 0, 0, 0))
               + chunk(b"IDAT", zlib.compress(b"\x00" + b"\xff\x00\x00" * 640))
               + chunk(b"IEND", b""))
        self.assertEqual(_png_size(png), (640, 480))
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "img.docx")
            doc = Document()
            doc.image(png, caption="cap")
            doc.save(out)
            _text, parts = _docx_text(out)
            import zipfile
            with zipfile.ZipFile(out) as z:
                rels = z.read("word/_rels/document.xml.rels").decode()
                body = z.read("word/document.xml").decode()
        self.assertIn("word/media/image1.png", parts)
        self.assertIn("/image", rels)
        self.assertIn("r:embed", body)




class WriteupTest(unittest.TestCase):
    def _hosts(self):
        from recce.models import Vuln
        h1 = Host(ip="10.0.20.5", hostnames=["web01"],
                  ports=[Port(portid=443, service="https")],
                  vulns=[Vuln(ip="10.0.20.5", port=443, protocol="tcp",
                              script_id="ssl-enum-ciphers",
                              title="Weak SSL/TLS ciphers or protocols",
                              severity="medium", source="config",
                              cwes=["CWE-327"], remediation="Disable weak ciphers.",
                              output="TLSv1.0 offered")])
        h2 = Host(ip="10.0.20.9", hostnames=["web02"],
                  ports=[Port(portid=443, service="https")],
                  vulns=[Vuln(ip="10.0.20.9", port=443, protocol="tcp",
                              script_id="ssl-enum-ciphers",
                              title="Weak SSL/TLS ciphers or protocols",
                              severity="medium", source="config", cwes=["CWE-327"],
                              output="RC4 offered"),
                         Vuln(ip="10.0.20.9", port=21, protocol="tcp",
                              script_id="version-db",
                              title="vsftpd 2.3.4 backdoor", severity="critical",
                              source="version-db", ids=["CVE-2011-2523"],
                              cwes=["CWE-78"], remediation="Upgrade vsftpd.")])
        return [h1, h2]

    def test_grouping_across_hosts(self):
        from recce.report_docx import group_findings
        findings = group_findings(self._hosts())
        # 2 distinct findings; critical sorts first.
        self.assertEqual([f.severity for f in findings], ["critical", "medium"])
        tls = next(f for f in findings if "SSL" in f.title)
        self.assertEqual(sorted(a[0] for a in tls.affected),
                         ["10.0.20.5", "10.0.20.9"])   # spans both hosts

    def test_build_writeups_and_no_overwrite(self):
        from recce.report_docx import build_writeups
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "writeups")
            summary = build_writeups(self._hosts(), out)
            self.assertEqual(summary["total"], 2)
            self.assertEqual(len(summary["written"]), 2)
            f_crit = next(p for p in os.listdir(out) if p.startswith("F-001"))
            text, _ = _docx_text(os.path.join(out, f_crit))
            for expect in ("F-001", "Affected systems:", "CWE-78",
                           "CVE-2011-2523", "Recommendations", "Evidence",
                           "Mission Risk and Impact", "[TESTER:"):
                self.assertIn(expect, text)
            # Re-run: existing files are kept, not overwritten.
            again = build_writeups(self._hosts(), out)
            self.assertEqual(len(again["written"]), 0)
            self.assertEqual(len(again["skipped"]), 2)

    def test_min_severity_filter(self):
        from recce.report_docx import build_writeups
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "w")
            summary = build_writeups(self._hosts(), out, min_severity="high")
            self.assertEqual(summary["total"], 1)   # only the critical

    def _hosts_potential_and_loot(self):
        from recce.models import Vuln
        return [Host(ip="10.0.30.5", hostnames=["box"],
                     ports=[Port(portid=23, service="telnet"),
                            Port(portid=445, service="microsoft-ds")],
                     local_findings=[{"section": "Sudo", "category": "sudo",
                                      "vector": "NOPASSWD find",
                                      "text": "NOPASSWD sudo: /usr/bin/find",
                                      "source": "recce-enum"}],
                     vulns=[
                         Vuln(ip="10.0.30.5", port=23, protocol="tcp",
                              script_id="version-db", title="Telnet cleartext",
                              severity="medium", source="version-db",
                              confidence="potential", cwes=["CWE-319"]),
                         Vuln(ip="10.0.30.5", port=445, protocol="tcp",
                              script_id="smb-vuln-ms17-010", title="smb-vuln-ms17-010",
                              severity="high", source="nse", ids=["CVE-2017-0143"],
                              output="VULNERABLE"),
                     ])]

    def test_potential_excluded_by_default_included_on_flag(self):
        from recce.report_docx import build_writeups
        hosts = self._hosts_potential_and_loot()
        with tempfile.TemporaryDirectory() as d:
            real = build_writeups(hosts, os.path.join(d, "r"))
            self.assertEqual(real["total"], 1)                 # only the nse ms17-010
            self.assertEqual(real["dropped_potential"], 1)     # telnet guess skipped
            allf = build_writeups(hosts, os.path.join(d, "a"), include_potential=True)
            self.assertEqual(allf["total"], 2)                 # both

    def test_list_findings_flags_real(self):
        from recce.report_docx import list_findings
        rows = list_findings(self._hosts_potential_and_loot())
        by_title = {r["title"]: r for r in rows}
        self.assertFalse(by_title["Telnet cleartext"]["real"])
        self.assertTrue(by_title["smb-vuln-ms17-010"]["real"])
        # stable ids: high sorts before the medium
        self.assertEqual(by_title["smb-vuln-ms17-010"]["id"], "F-001")

    def test_blank_confidence_reads_reported_everywhere(self):
        # Regression: list_findings coerced a blank confidence to "confirmed", so the
        # HTML summary table showed a green "Confirmed" badge for a finding the detail
        # card and the exec "Confirmed" tile both treat as "Reported" - an honesty
        # inconsistency. A blank confidence (the NSE-VULNERABLE ms17-010 here) must
        # read the same in all three places.
        from recce.report_docx import list_findings, group_findings
        from recce import report_html as rh
        hosts = self._hosts_potential_and_loot()
        row = {r["title"]: r for r in list_findings(hosts)}["smb-vuln-ms17-010"]
        self.assertEqual(row["confidence"], "")        # not coerced to "confirmed"
        # Summary-table badge (from the dict) matches the detail-card badge (from the
        # grouped Finding), and neither claims "Confirmed".
        f = {g.title: g for g in group_findings(hosts)}["smb-vuln-ms17-010"]
        self.assertEqual(rh._conf_badge(row["confidence"]), rh._conf_badge(f.confidence))
        self.assertIn("Reported", rh._conf_badge(row["confidence"]))
        self.assertNotIn("Confirmed", rh._conf_badge(row["confidence"]))
        # The exec-summary "Confirmed" tile counts none of the blank-confidence rows,
        # so its number can't exceed the "Confirmed" badges in the table.
        confirmed = sum(1 for g in group_findings(hosts)
                        if (g.confidence or "").lower() == "confirmed")
        self.assertEqual(confirmed, 0)

    def test_single_writeup_prefills_looted(self):
        from recce.report_docx import build_one_writeup
        hosts = self._hosts_potential_and_loot()
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "writeups")
            res = build_one_writeup(hosts, out, "ms17")
            self.assertTrue(res["written"])
            self.assertEqual(res["looted"], 1)
            text, _ = _docx_text(res["written"])
            self.assertIn("Obtained Access / Looted Evidence", text)
            self.assertIn("NOPASSWD sudo: /usr/bin/find", text)

    def test_single_writeup_selectors(self):
        from recce.report_docx import build_one_writeup
        hosts = self._hosts_potential_and_loot()
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "w")
            # by F-id
            self.assertTrue(build_one_writeup(hosts, out, "F-001")["written"])
            # by CVE
            self.assertTrue(build_one_writeup(hosts, out, "CVE-2017-0143",
                                              overwrite=True)["written"])
            # ambiguous IP -> lists candidates, writes nothing
            amb = build_one_writeup(hosts, out, "10.0.30.5")
            self.assertIsNone(amb["written"])
            self.assertEqual(amb["reason"], "ambiguous")
            self.assertEqual(len(amb["matched"]), 2)
            # unmatched
            none = build_one_writeup(hosts, out, "zzz-nope")
            self.assertIsNone(none["written"])
            self.assertEqual(none["reason"], "none")

    def test_auto_walkthrough_steps(self):
        from recce.report_docx import group_findings, _walkthrough_steps
        findings = group_findings(self._hosts())
        tls = next(f for f in findings if "SSL" in f.title)
        steps = _walkthrough_steps(tls)
        self.assertTrue(steps)
        joined = " ".join(steps)
        self.assertIn("nmap -sV", joined)         # discovery step
        self.assertIn("ssl-enum-ciphers", joined)  # tailored confirmation step

    def test_narrative_is_multi_paragraph_and_context_aware(self):
        from recce.models import Vuln
        from recce.report_docx import group_findings, _narrative
        # A likely (version-matched) web finding.
        web = Host(ip="10.0.20.5", hostnames=["web01"],
                   ports=[Port(portid=80, service="http", product="Apache httpd",
                               version="2.4.49")],
                   vulns=[Vuln(ip="10.0.20.5", port=80, protocol="tcp",
                               script_id="version-db", title="Apache path traversal",
                               severity="critical", source="version-db",
                               confidence="likely", cwes=["CWE-22"],
                               ids=["CVE-2021-41773"])])
        f = group_findings([web])[0]
        paras = _narrative(f)
        self.assertEqual(len(paras), 3)                       # context / finding / impact
        blob = " ".join(paras).lower()
        self.assertIn("web service", blob)                    # service context
        self.assertIn("apache httpd 2.4.49", blob)            # detected product
        self.assertIn("10.0.20.5", blob)                      # affected host named
        self.assertIn("cve-2021-41773", blob)                 # CVE woven in
        self.assertIn("critical-risk", blob)                  # severity framing
        self.assertIn("read files outside", blob)             # CWE-22 plain impact
        self.assertIn("range known to be affected", blob)     # likely-confidence note

        # A potential (advisory) finding gets the "confirm by hand" caveat instead.
        adv = Host(ip="10.0.10.10", hostnames=["dc01"], os_family="Windows",
                   ports=[Port(portid=445, service="microsoft-ds",
                               product="Windows Server 2019")],
                   vulns=[Vuln(ip="10.0.10.10", port=445, protocol="tcp",
                               script_id="version-db", title="verify ZeroLogon",
                               severity="critical", source="version-db",
                               confidence="potential", cwes=["CWE-330"])])
        pa = " ".join(_narrative(group_findings([adv])[0])).lower()
        self.assertIn("smb", pa)                              # SMB service context
        self.assertIn("confirmed through hands-on", pa)       # potential caveat

    def test_every_cwe_is_classified_named_and_has_an_impact(self):
        """Guarantee: every CWE recce can emit maps to a type + a name, and every
        type has a plain-language impact - so no finding drops to a blank type."""
        import glob
        import re
        from recce.report_docx import _CWE_TYPE, _CWE_NAME, _TYPE_IMPACT
        used = set()
        for fn in glob.glob(os.path.join(os.path.dirname(os.path.dirname(
                os.path.abspath(__file__))), "recce", "*.py")):
            # report_docx.py and cwe.py are the CWE naming/typing tables themselves - they
            # LIST CWEs for reference, they don't EMIT them as findings, so they aren't
            # bound by the "every emitted CWE must be typed here" guarantee.
            if fn.endswith(("report_docx.py", "cwe.py")):
                continue
            with open(fn) as fh:
                used |= set(re.findall(r"CWE-\d+", fh.read()))
        self.assertTrue(used)
        typed = set()
        for keys, _label, _cia in _CWE_TYPE:
            typed |= set(keys)
        self.assertEqual(used - typed, set(), "CWEs with no vulnerability type")
        self.assertEqual(used - set(_CWE_NAME), set(), "CWEs with no reference name")
        for _keys, label, _cia in _CWE_TYPE:
            self.assertIn(label, _TYPE_IMPACT, f"type '{label}' has no impact wording")
        # CWEs the NSE-script mapper can assign must also be named + typed.
        from recce.report_docx import _NSE_CWE
        nse_cwes = {c for cs in _NSE_CWE.values() for c in cs}
        self.assertEqual(nse_cwes - typed, set(), "NSE-mapped CWEs with no type")
        self.assertEqual(nse_cwes - set(_CWE_NAME), set(), "NSE-mapped CWEs with no name")

    def test_nse_scripts_auto_map_to_cwe_and_cve(self):
        from recce.models import Vuln
        from recce.report_docx import group_findings

        def finding_for(script_id, title=None):
            h = Host(ip="10.0.0.9", ports=[Port(portid=445, service="microsoft-ds")],
                     vulns=[Vuln(ip="10.0.0.9", port=445, protocol="tcp",
                                 script_id=script_id, title=title or script_id,
                                 severity="high", source="nse")])
            return group_findings([h])[0]

        # ms17-010 (no CVE in the id) -> mapped CVE + CWE.
        f = finding_for("smb-vuln-ms17-010")
        self.assertIn("CVE-2017-0144", f.cves)
        self.assertIn("CWE-787", f.cwes)
        # http-vuln-cveYYYY-N -> CVE parsed from the id + CWE mapped.
        f = finding_for("http-vuln-cve2021-41773")
        self.assertIn("CVE-2021-41773", f.cves)
        self.assertIn("CWE-22", f.cwes)
        # Heartbleed TLS script -> its CVE + CWE.
        f = finding_for("ssl-heartbleed")
        self.assertIn("CVE-2014-0160", f.cves)
        self.assertIn("CWE-125", f.cwes)
        # A version-db finding that already has CWE/CVE is NOT overridden.
        h = Host(ip="10.0.0.9", ports=[Port(portid=80, service="http")],
                 vulns=[Vuln(ip="10.0.0.9", port=80, protocol="tcp",
                             script_id="version-db", title="Apache thing",
                             severity="high", source="version-db",
                             cwes=["CWE-22"], ids=["CVE-2021-41773"])])
        f = group_findings([h])[0]
        self.assertEqual(f.cwes, ["CWE-22"])

    def test_marquee_vulns_get_specific_impact(self):
        from recce.models import Vuln
        from recce.report_docx import group_findings, _narrative
        cases = [
            (["CVE-2020-1472"], "verify zerologon", "ZeroLogon"),
            (["CVE-2021-34527"], "printnightmare", "Print Spooler"),
            ([], "smb-vuln-ms17-010", "EternalBlue"),        # NSE hit, no CVE
            (["CVE-2020-0796"], "smbghost", "SMBv3"),
        ]
        for cves, title, needle in cases:
            h = Host(ip="10.0.0.9", os_family="Windows",
                     ports=[Port(portid=445, service="microsoft-ds")],
                     vulns=[Vuln(ip="10.0.0.9", port=445, protocol="tcp",
                                 script_id=title, title=title, severity="critical",
                                 source="nse", ids=cves)])
            blob = " ".join(_narrative(group_findings([h])[0]))
            self.assertIn(needle, blob, f"{title} missing marquee wording")

    def test_reports_exclude_informational_by_default(self):
        from recce.models import Vuln
        from recce.report_docx import build_writeups
        h = Host(ip="10.0.0.9", ports=[Port(portid=25, service="smtp")],
                 vulns=[
                     Vuln(ip="10.0.0.9", port=25, protocol="tcp", script_id="a",
                          title="SMTP server exposed", severity="info", source="version-db"),
                     Vuln(ip="10.0.0.9", port=25, protocol="tcp", script_id="b",
                          title="Weak TLS on SMTP", severity="medium", source="probe"),
                 ])
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "w")
            summary = build_writeups([h], out)               # default = findings only
            self.assertEqual(summary["total"], 1)            # the medium, not the info
            names = os.listdir(out)
            self.assertFalse(any("_info_" in n for n in names))
            # Opting in re-includes informational items.
            summary2 = build_writeups([h], os.path.join(d, "w2"), min_severity="info")
            self.assertEqual(summary2["total"], 2)

    def test_walkthrough_uses_searchsploit_exploit(self):
        from recce.models import Vuln, Exploit
        from recce.report_docx import group_findings, _walkthrough_steps
        h = Host(ip="10.0.20.6", ports=[Port(portid=21, service="ftp",
                 product="vsftpd", version="2.3.4")],
                 vulns=[Vuln(ip="10.0.20.6", port=21, protocol="tcp",
                             script_id="version-db", title="vsftpd 2.3.4 backdoor",
                             severity="critical", source="version-db",
                             ids=["CVE-2011-2523"])],
                 exploits=[Exploit(ip="10.0.20.6", port=21, edb_id="17491",
                                   title="vsftpd 2.3.4 backdoor")])
        f = group_findings([h])[0]
        self.assertIn("17491", " ".join(_walkthrough_steps(f)))

    def test_walkthrough_only_cites_proven_exploits(self):
        from recce.models import Vuln
        from recce.report_docx import group_findings, _walkthrough_steps

        def steps(title, conf, cves, source="version-db", svc="http", port=80):
            h = Host(ip="1.1.1.1", ports=[Port(portid=port, service=svc)],
                     vulns=[Vuln(ip="1.1.1.1", port=port, protocol="tcp",
                                 script_id=source, title=title, severity="high",
                                 source=source, confidence=conf, ids=cves)])
            return " ".join(_walkthrough_steps(group_findings([h])[0]))

        # Proven exploit (curated) on a version-matched finding -> cited concretely.
        s = steps("Apache path traversal", "likely", ["CVE-2021-41773"])
        self.assertIn("Metasploit", s)
        self.assertIn("apache_normalize_path_rce", s)
        # NSE-confirmed ms17-010 -> proven EternalBlue exploit cited.
        self.assertIn("eternalblue", steps("smb-vuln-ms17-010", "", [],
                                            source="nse", svc="microsoft-ds", port=445).lower())
        # Advisory/potential finding -> NO exploit line, even with a famous CVE.
        s = steps("Windows DC - verify ZeroLogon", "potential", ["CVE-2020-1472"],
                  svc="microsoft-ds", port=445)
        self.assertNotIn("Metasploit", s)
        self.assertNotIn("exploit", s.lower())
        # A version match with no proven exploit known -> no speculative "research" line.
        s = steps("OpenSSH username enumeration", "likely", ["CVE-2018-15473"],
                  svc="ssh", port=22)
        self.assertNotIn("Metasploit", s)
        self.assertNotIn("Research a working exploit", s)

    def test_combined_report(self):
        from recce.report_docx import build_combined
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "combined.docx")
            res = build_combined(self._hosts(), out, title="Test Engagement")
            self.assertEqual(res["total"], 2)
            text, parts = _docx_text(out)
            import zipfile
            with zipfile.ZipFile(out) as z:
                body = z.read("word/document.xml").decode()
                settings = z.read("word/settings.xml").decode()
                ctypes = z.read("[Content_Types].xml").decode()
            self.assertIn("<w:tbl>", body)              # has tables
            self.assertIn("Test Engagement", text)      # title
            self.assertIn("Summary", text)
            self.assertIn("F-001", text)                # findings numbered
            self.assertIn("vsftpd 2.3.4 backdoor", text)
            # Auto-updating table of contents: the TOC field + the updateFields
            # setting that makes Word rebuild it on open (jump-list over findings).
            self.assertIn("Contents", text)
            self.assertIn('w:fldCharType="begin"', body)
            self.assertIn('TOC \\o', body)
            self.assertIn('<w:updateFields w:val="true"/>', settings)
            self.assertIn("/word/settings.xml", ctypes)   # declared, so Word reads it
            # Severity summary counts carry the ramp colour (critical = red).
            self.assertIn('w:color w:val="C00000"', body)

    def test_screenshot_url_classification(self):
        from recce import screenshot
        self.assertTrue(screenshot._web_url(Port(portid=443, service="https")))
        self.assertTrue(screenshot._web_url(Port(portid=8080, service="http-proxy")))
        self.assertIsNone(screenshot._web_url(Port(portid=22, service="ssh")))
        # No browser in the test env -> capture is a no-op, never raises.
        h = Host(ip="1.2.3.4", ports=[Port(portid=80, service="http")])
        if not screenshot.available():
            self.assertEqual(screenshot.capture_for_host(h), [])

    def _fake_browser(self, name):
        """Create a fake executable and point RECCE_BROWSER at it."""
        d = tempfile.mkdtemp()
        self.addCleanup(lambda: __import__("shutil").rmtree(d, ignore_errors=True))
        path = os.path.join(d, name)
        with open(path, "w") as fh:
            fh.write("#!/bin/sh\n")
        os.chmod(path, 0o755)
        return path

    def test_browser_found_off_path(self):
        """Regression: a browser installed but not on PATH (sudo secure_path,
        snap, /opt) must still be found via the absolute-path fallback scan."""
        import shutil as _sh
        from recce import screenshot as s
        d = tempfile.mkdtemp()
        self.addCleanup(lambda: _sh.rmtree(d, ignore_errors=True))
        # a browser in a bin dir + one nested under an /opt-style dir
        bind = os.path.join(d, "bin"); os.makedirs(bind)
        optd = os.path.join(d, "opt", "vendor"); os.makedirs(optd)
        chromium = os.path.join(bind, "chromium")
        firefox = os.path.join(optd, "firefox")
        for p in (chromium, firefox):
            with open(p, "w") as fh:
                fh.write("#!/bin/sh\n")
            os.chmod(p, 0o755)
        orig_dirs, orig_globs = s._SCAN_DIRS, s._OPT_GLOBS
        orig_path = os.environ.get("PATH", "")
        os.environ.pop("RECCE_BROWSER", None)
        try:
            os.environ["PATH"] = "/nonexistent-xyz"   # nothing resolvable on PATH
            # 1) scan-dir fallback finds the bin-dir chromium
            s._SCAN_DIRS = [bind]; s._OPT_GLOBS = []
            self.assertEqual(s.browser_tool(), chromium)
            self.assertTrue(s.available())
            # 2) /opt-style glob finds the nested firefox
            s._SCAN_DIRS = []
            s._OPT_GLOBS = [os.path.join(d, "opt", "*/{n}")]
            self.assertEqual(s.browser_tool(), firefox)
        finally:
            s._SCAN_DIRS, s._OPT_GLOBS = orig_dirs, orig_globs
            os.environ["PATH"] = orig_path

    def test_firefox_detection_and_command(self):
        from recce import screenshot
        ff = self._fake_browser("firefox")
        os.environ["RECCE_BROWSER"] = ff
        self.addCleanup(lambda: os.environ.pop("RECCE_BROWSER", None))
        try:
            self.assertEqual(screenshot.browser_tool(), ff)
            self.assertTrue(screenshot._is_firefox(ff))
            self.assertTrue(screenshot.available())

            captured = {}

            def fake_run(cmd, **kw):
                captured["cmd"] = cmd
                # Emulate Firefox writing the screenshot: -screenshot <out> URL
                out = cmd[cmd.index("--screenshot") + 1]
                with open(out, "wb") as fh:
                    fh.write(b"\x89PNG\r\n\x1a\n" + b"\0" * 32)
                return SimpleNamespace(returncode=0, stdout=b"", stderr=b"")

            import subprocess as _sp
            orig = _sp.run
            _sp.run = fake_run
            try:
                png = screenshot.capture("http://1.2.3.4:80/")
            finally:
                _sp.run = orig

            self.assertIsNotNone(png)
            self.assertTrue(png.startswith(b"\x89PNG"))
            cmd = captured["cmd"]
            self.assertEqual(os.path.basename(cmd[0]), "firefox")
            self.assertIn("--headless", cmd)
            self.assertIn("-profile", cmd)
            # Screenshot path is a positional arg (no `=` form), URL is last.
            self.assertEqual(cmd[-1], "http://1.2.3.4:80/")
            self.assertNotIn("--ignore-certificate-errors", cmd)
        finally:
            os.environ.pop("RECCE_BROWSER", None)

    def test_chrome_detection_and_command(self):
        from recce import screenshot
        ch = self._fake_browser("chromium")
        os.environ["RECCE_BROWSER"] = ch
        self.addCleanup(lambda: os.environ.pop("RECCE_BROWSER", None))
        try:
            self.assertFalse(screenshot._is_firefox(ch))
            captured = {}

            def fake_run(cmd, **kw):
                captured["cmd"] = cmd
                out = cmd[-2].split("=", 1)[1]
                with open(out, "wb") as fh:
                    fh.write(b"\x89PNG\r\n\x1a\n" + b"\0" * 32)
                return SimpleNamespace(returncode=0, stdout=b"", stderr=b"")

            import subprocess as _sp
            orig = _sp.run
            _sp.run = fake_run
            try:
                png = screenshot.capture("https://1.2.3.4:443/")
            finally:
                _sp.run = orig

            self.assertIsNotNone(png)
            cmd = captured["cmd"]
            self.assertIn("--headless", cmd)
            self.assertIn("--ignore-certificate-errors", cmd)
            self.assertTrue(cmd[-2].startswith("--screenshot="))
            self.assertEqual(cmd[-1], "https://1.2.3.4:443/")
        finally:
            os.environ.pop("RECCE_BROWSER", None)




class StepCheckboxTest(unittest.TestCase):
    def _host(self, **kw):
        h = Host(ip="10.0.0.5", subnet="10.0.0.0/24",
                 ports=[Port(portid=80, service="http"), Port(portid=3306, service="mysql")])
        for k, v in kw.items():
            setattr(h, k, v)
        return h

    def test_step_auto(self):
        h = self._host()
        self.assertFalse(tr.step_auto(h, "enum"))
        h.enumerated = True
        self.assertTrue(tr.step_auto(h, "enum"))
        self.assertFalse(tr.step_auto(h, "vuln"))   # ports not scanned
        for p in h.ports:
            p.vuln_scanned = True
        self.assertTrue(tr.step_auto(h, "vuln"))
        self.assertFalse(tr.step_auto(h, "db"))     # has mysql, not db_scanned
        h.db_scanned = True
        self.assertTrue(tr.step_auto(h, "db"))

    def test_db_not_applicable_when_no_db(self):
        # An SSH-only Linux host: DB, Web and AD steps simply don't apply.
        h = Host(ip="10.0.0.6", os_family="Linux", enumerated=True,
                 ports=[Port(portid=22, service="ssh")])
        self.assertFalse(tr.step_applies(h, "db"))
        self.assertFalse(tr.step_applies(h, "web"))
        self.assertFalse(tr.step_applies(h, "ad"))
        # Universal steps still apply.
        self.assertTrue(tr.step_applies(h, "enum"))
        self.assertTrue(tr.step_applies(h, "vuln"))

    def test_step_applicability_by_surface(self):
        web = Host(ip="10.0.0.7", os_family="Linux",
                   ports=[Port(portid=443, service="https")])
        self.assertTrue(tr.step_applies(web, "web"))
        self.assertFalse(tr.step_applies(web, "ad"))   # Linux web, not a DC
        self.assertFalse(tr.step_applies(web, "db"))

        # A plain SMB file server is NOT an AD host (SMB is tracked on Services).
        smb = Host(ip="10.0.0.8", os_family="Windows",
                   ports=[Port(portid=445, service="microsoft-ds")])
        self.assertFalse(tr.step_applies(smb, "ad"))

        # A DC (LDAP/Kerberos) is an AD host.
        dc = Host(ip="10.0.0.10", os_family="Windows",
                  ports=[Port(portid=389, service="ldap"),
                         Port(portid=88, service="kerberos-sec")])
        self.assertTrue(tr.step_applies(dc, "ad"))

        # Kill-chain markers apply to anything with an open port.
        for step in ("access", "creds", "lateral"):
            self.assertTrue(tr.step_applies(web, step))
        dead = Host(ip="10.0.0.11", state="up", ports=[])
        for step in ("access", "creds", "lateral", "vuln"):
            self.assertFalse(tr.step_applies(dead, step))

        # Priv-esc only applies once the phase has run (a foothold exists).
        self.assertFalse(tr.step_applies(dc, "privesc"))
        dc.privesc_checked = True
        self.assertTrue(tr.step_applies(dc, "privesc"))

    def test_manual_steps_never_auto_check(self):
        # AD review + kill-chain markers are operator sign-offs: applicable but
        # never auto-completed by the tool, even after enumeration.
        dc = Host(ip="10.0.0.10", os_family="Windows", enumerated=True,
                  roles=["Domain Controller"],
                  ports=[Port(portid=389, service="ldap"),
                         Port(portid=88, service="kerberos-sec")])
        for step in ("ad", "access", "creds", "lateral"):
            self.assertTrue(tr.step_applies(dc, step))
            self.assertFalse(tr.step_auto(dc, step))

    def test_manual_marker_ticks_persist(self):
        # Ticking a manual kill-chain box is recorded as an override and, unlike
        # auto steps, no phase clears it.
        from recce import cli
        with tempfile.TemporaryDirectory() as d:
            store = Store(os.path.join(d, "t.sqlite"))
            store.upsert_host(Host(ip="10.0.0.5", subnet="10.0.0.0/24",
                                   enumerated=True,
                                   ports=[Port(portid=80, service="http")]))
            akey = tr.step_key("access", "10.0.0.5")
            cli._reconcile_steps(store, {akey: (True, "")})   # tester ticked it
            self.assertTrue(store.get_tracking()[akey][0])
            # Unticking matches the auto default (False) -> override cleared.
            cli._reconcile_steps(store, {akey: (False, "")})
            self.assertNotIn(akey, store.get_tracking())
            store.close()

    def test_web_step_auto_done_when_web_ports_scanned(self):
        h = Host(ip="10.0.0.9", enumerated=True,
                 ports=[Port(portid=80, service="http"), Port(portid=22, service="ssh")])
        self.assertFalse(tr.step_auto(h, "web"))
        h.ports[0].vuln_scanned = True    # the web port got probed
        self.assertTrue(tr.step_auto(h, "web"))

    def test_na_step_renders_dash_and_no_override(self):
        # A Linux SSH box: the DB/Web/AD columns show N/A, not a checkbox, and
        # reading the workbook back records no step override for them.
        h = Host(ip="10.0.0.6", os_family="Linux", enumerated=True,
                 ports=[Port(portid=22, service="ssh")])
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook([h], out)
            rows = xlsx.read_sheets(out)["Checklist"]
            hidx = header_index(rows, "IP")
            header = rows[hidx]
            # The row after the header is the collapsible subnet band; take the first host row.
            row = next(r for r in rows[hidx + 1:] if r[0] in (xlsx.CHECK_ON, xlsx.CHECK_OFF))
            for col in ("DB", "Web", "AD"):
                self.assertEqual(row[header.index(col)], tr.STEP_NA)
            back = read_workbook_tracking(out)
            self.assertNotIn(tr.step_key("db", "10.0.0.6"), back)
            self.assertNotIn(tr.step_key("web", "10.0.0.6"), back)
            self.assertNotIn(tr.step_key("ad", "10.0.0.6"), back)
            # Universal steps (enum + kill-chain, host has an open port) are tracked.
            self.assertIn(tr.step_key("enum", "10.0.0.6"), back)
            self.assertIn(tr.step_key("access", "10.0.0.6"), back)

    def test_checkbox_reflects_auto_then_override(self):
        h = self._host(enumerated=True)
        for p in h.ports:
            p.vuln_scanned = True
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            # No override -> follows auto (vuln done -> TRUE).
            build_workbook([h], out)
            back = read_workbook_tracking(out)
            self.assertTrue(back[tr.step_key("vuln", "10.0.0.5")][0])
            # Override FALSE -> checkbox shows FALSE despite auto TRUE.
            build_workbook([h], out, tracking={tr.step_key("vuln", "10.0.0.5"): (False, "")})
            back = read_workbook_tracking(out)
            self.assertFalse(back[tr.step_key("vuln", "10.0.0.5")][0])

    def test_services_vulnscan_not_read_as_step(self):
        # The Services sheet also has a "Vuln-scan" column; it must NOT pollute steps.
        h = self._host(enumerated=True)  # ports NOT vuln_scanned -> Services shows "pending"
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "wb.xlsx")
            build_workbook([h], out)
            back = read_workbook_tracking(out)
        # vuln step comes from the Checklist (auto = pending = False), and there's
        # exactly one value - not overwritten to False by the Services rows.
        self.assertIn(tr.step_key("vuln", "10.0.0.5"), back)

    def test_reconcile_records_and_clears_override(self):
        from recce import cli
        with tempfile.TemporaryDirectory() as d:
            store = Store(os.path.join(d, "t.sqlite"))
            h = self._host(enumerated=True)
            for p in h.ports:
                p.vuln_scanned = True     # vuln auto = True
            store.upsert_host(h)
            key = tr.step_key("vuln", "10.0.0.5")
            # Shown FALSE but auto TRUE -> record override.
            cli._reconcile_steps(store, {key: (False, "")})
            self.assertIn(key, store.get_tracking())
            # Shown TRUE matches auto -> clear override.
            cli._reconcile_steps(store, {key: (True, "")})
            self.assertNotIn(key, store.get_tracking())
            store.close()




class CheckboxPersistenceTest(unittest.TestCase):
    def test_every_checkbox_header_round_trips(self):
        """Every column with the checkbox role must be recognised by the read-back
        (CHECKBOX_HEADERS), or the operator's ticks are silently lost on regen."""
        from recce import report_excel as rx
        from recce.models import Vuln, Credential
        hosts = [Host(ip="10.0.0.5", os_family="Windows", roles=["Domain Controller"],
                      local_findings=[{"category": "sudo",
                                       "vector": "NOPASSWD sudo: /usr/bin/find",
                                       "section": "Sudo", "source": "recce-enum"}],
                      accounts=[__import__("recce.models", fromlist=["Account"]).Account(
                          ip="10.0.0.5", source="nse", kind="domain", domain="CORP")],
                      ports=[Port(portid=445, service="microsoft-ds")],
                      vulns=[Vuln(ip="10.0.0.5", port=445, protocol="tcp",
                                  script_id="smb-vuln-ms17-010", title="ms17-010",
                                  severity="high", source="nse", ids=["CVE-2017-0143"],
                                  output="VULNERABLE")])]
        creds = [Credential(username="alice", secret="Pw!", domain="CORP")]
        pre, ad_specs, tail = rx._ordered_specs(hosts, None, creds)
        for spec in pre + list(ad_specs.values()) + tail:
            cb = [h for h, role, _w in spec.cols if role == "checkbox"]
            for header in cb:
                self.assertIn(header, rx.CHECKBOX_HEADERS,
                              f"{spec.title}: checkbox column {header!r} not in "
                              "CHECKBOX_HEADERS -> ticks won't persist")




class HtmlReportTest(unittest.TestCase):
    def _hosts(self):
        from recce.models import Vuln
        return [Host(ip="10.0.0.5", hostnames=["dc01"], os_family="Windows",
                     roles=["Domain Controller"], defenses=["EDR/AV: CSFalcon (process)"],
                     ports=[Port(portid=445, service="microsoft-ds")],
                     vulns=[Vuln(ip="10.0.0.5", port=445, protocol="tcp",
                                 script_id="smb-vuln-ms17-010",
                                 title="smb-vuln-ms17-010 <x>", severity="high",
                                 source="nse", ids=["CVE-2017-0143"],
                                 output="VULNERABLE")])]

    def test_self_contained_and_escaped(self):
        from recce import report_html
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "report.html")
            report_html.build_html(self._hosts(), p, title="Client X",
                                   generated="2026-01-01")
            with open(p, encoding="utf-8") as fh:
                html = fh.read()
        self.assertIn("<!doctype html>", html)
        # self-contained: no external resources at all.
        for bad in ("http://", "https://", "src=", "<link"):
            self.assertNotIn(bad, html)
        self.assertIn("Client X", html)
        for section in ("Executive summary", "Findings by severity", "Attack path",
                        "Hosts", "CVE-2017-0143"):
            self.assertIn(section, html)
        self.assertIn("smb-vuln-ms17-010 &lt;x&gt;", html)   # HTML-escaped title
        # Attack-path graph is embedded as inline SVG — renders offline, no tools/JS.
        self.assertIn('aria-label="Attack path"', html)
        self.assertIn("<svg", html)
        # The attack path is framed honestly: projected, precondition-grounded, and
        # explicitly NOT executed by recce (nothing reads as a proven kill chain).
        self.assertIn("projected", html)
        self.assertIn("not</b> been walked end-to-end", html)
        self.assertIn("recce does not exploit", html)

    def test_print_css_keeps_findings_whole_and_unclips_evidence(self):
        """The report is routinely delivered as PDF: a finding card must not split
        across a page break, and evidence must print in full (not clip at the on-screen
        max-height). Locks the print-quality contract against a CSS refactor."""
        import re
        from recce import report_html
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "r.html")
            report_html.build_html(self._hosts(), p, title="Print")
            html = open(p, encoding="utf-8").read()
        m = re.search(r"@media print\{.*?\n\}", html, re.S)
        self.assertIsNotNone(m, "no @media print block")
        block = m.group(0)
        # A finding card, an attack-path step, and the diagram stay whole; rows never split.
        for cls in (".fcard", ".step", ".stage", ".netmap", "tr"):
            self.assertIn(cls, block, f"{cls} not protected from a page break in print")
        self.assertIn("break-inside:avoid", block)
        self.assertIn("max-height:none", block)           # evidence not clipped in a PDF
        self.assertIn(".netmap svg{max-width:100%", block)  # wide map fits the page

    def test_detailed_findings_section(self):
        from recce import report_html
        from recce.models import Vuln
        h = Host(ip="10.0.0.6", os_family="Linux",
                 ports=[Port(portid=21, service="ftp")],
                 vulns=[Vuln(ip="10.0.0.6", port=21, protocol="tcp",
                             script_id="vsftpd-backdoor", title="vsFTPd backdoor",
                             severity="critical", source="nse",
                             ids=["CVE-2011-2523"],
                             remediation="Upgrade vsFTPd to a patched release.",
                             output="Backdoor shell on 6200 confirmed")])
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "r.html")
            report_html.build_html([h], p)
            with open(p, encoding="utf-8") as fh:
                html = fh.read()
        self.assertIn("Finding details", html)
        self.assertIn("class=\"fcard\"", html)
        self.assertIn("Upgrade vsFTPd", html)                 # remediation card
        self.assertIn("Backdoor shell on 6200 confirmed", html)  # evidence excerpt
        self.assertIn("10.0.0.6:21", html)                    # affected system

    def test_visual_dashboard(self):
        """The 'At a glance' dashboard renders inline-SVG visuals for a non-technical
        reader, and stays self-contained (no xmlns/external refs in the SVG)."""
        from recce import report_html
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "report.html")
            report_html.build_html(self._hosts(), p, title="Viz")
            with open(p, encoding="utf-8") as fh:
                html = fh.read()
        for section in ("At a glance", "Findings by severity", "Machines by risk"):
            self.assertIn(section, html)
        self.assertIn("<svg", html)                      # the severity donut
        self.assertIn("<circle", html)                   # donut segment(s)
        self.assertIn('class="hbar"', html)              # risk / affected bars
        self.assertNotIn("xmlns", html)                  # inline SVG stays self-contained
        for bad in ("src=", "<link", "<script"):         # nothing is fetched
            self.assertNotIn(bad, html)

    def test_scoring_legend_and_grounding(self):
        """The report explains why a severity is assigned and never presents an
        unverified finding as fact: a scoring legend (severity bands + confidence
        meanings), a per-finding confidence badge + 'why this rating' basis line, and
        a grounded exec assessment."""
        from recce import report_html
        from recce.models import Vuln
        h = Host(ip="10.0.0.7", os_family="Linux",
                 ports=[Port(portid=80, service="http")],
                 vulns=[Vuln(ip="10.0.0.7", port=80, protocol="tcp",
                             script_id="apache-cve", title="Apache < 2.4.59 vulns",
                             severity="high", source="version-db",
                             confidence="potential", ids=["CVE-2024-27316"],
                             output="Apache httpd 2.4.41")])
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "r.html")
            report_html.build_html([h], p, title="Grounding")
            with open(p, encoding="utf-8") as fh:
                html = fh.read()
        # Scoring legend explains severity + confidence.
        self.assertIn("How findings are scored", html)
        self.assertIn("CVSS", html)
        for level in ("Critical", "High", "Medium", "Low"):
            self.assertIn(level, html)
        # A potential finding is labelled as such (not shown as fact) and the exec
        # assessment flags it for verification.
        self.assertIn("Potential", html)
        self.assertIn("flagged for manual verification", html)
        # Per-finding 'why this rating' basis references the CVSS/CVE source.
        self.assertIn('class="basis"', html)
        self.assertIn("CVE-2024-27316", html)

    def test_coverage_checklist_mirrors_tracking(self):
        """A read-only 'Assessment coverage' checklist reflects both the tool's auto
        state (enumerated host -> Enumerated done) and an operator tick passed via
        tracking (a reviewed host shows checked)."""
        from recce import report_html, tracking as tr
        h = Host(ip="10.0.0.10", subnet="10.0.0.0/24", state="up",
                 up_reason="syn-ack", enumerated=True,
                 ports=[Port(portid=445, state="open", service="microsoft-ds")])
        ticks = {tr.host_key("10.0.0.10"): (True, "done here")}
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "r.html")
            report_html.build_html([h], p, title="Cov", tracking=ticks)
            with open(p, encoding="utf-8") as fh:
                html = fh.read()
        self.assertIn("Assessment coverage", html)
        self.assertIn('table class="cov"', html)
        self.assertIn("10.0.0.10", html)
        self.assertIn("&#10003;", html)                  # at least one ✓ (enum/reviewed)
        self.assertIn("1/1 reviewed", html)              # the operator tick is mirrored
        # Still self-contained.
        for bad in ("src=", "<link", "<script"):
            self.assertNotIn(bad, html)

    def test_users_and_credentials_inventory(self):
        """All users and captured credentials are surfaced on the companion assets
        page; the credential secret is masked in the shareable HTML (full values stay
        in the workbook)."""
        from recce import report_html
        from recce.models import Account, Credential
        h = Host(ip="10.0.10.10", state="up", up_reason="syn-ack",
                 ports=[Port(portid=445, state="open", service="microsoft-ds")],
                 accounts=[
                     Account(ip="10.0.10.10", source="ldap", kind="user", name="jdoe",
                             domain="corp.local", rid="1104",
                             attrs={"admincount": "1", "description": "IT admin"}),
                     Account(ip="10.0.10.10", source="netexec", kind="share",
                             name="ADMIN$", detail="READ,WRITE")])
        creds = [Credential(username="admin", secret="Passw0rd!", kind="password",
                            domain="corp.local", source="secretsdump",
                            origin_ip="10.0.10.10")]
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "assets.html")
            report_html.build_assets_html([h], p, credentials=creds,
                                          report_link="report.html")
            with open(p, encoding="utf-8") as fh:
                html = fh.read()
        self.assertIn("Users &amp; accounts", html)
        self.assertIn("jdoe", html)
        self.assertIn("admin", html)                    # AdminCount flag pill
        self.assertIn("ADMIN$", html)                   # share listed
        self.assertIn("Credentials captured", html)
        self.assertIn("secretsdump", html)
        self.assertNotIn("Passw0rd!", html)             # secret is MASKED
        self.assertIn("[9 chars]", html)                # masked length shown
        self.assertIn("Findings report", html)          # cross-link back to report
        # The findings report itself no longer carries these sections.
        with tempfile.TemporaryDirectory() as d:
            rp = os.path.join(d, "report.html")
            report_html.build_html([h], rp, credentials=creds,
                                   assets_link="assets.html")
            with open(rp, encoding="utf-8") as fh:
                report = fh.read()
        self.assertNotIn("Credentials captured", report)
        self.assertNotIn("Users &amp; accounts", report)
        self.assertIn("Architecture &amp; assets", report)   # link to companion page

    def test_empty_hosts_ok(self):
        from recce import report_html
        with tempfile.TemporaryDirectory() as d:
            p = os.path.join(d, "r.html")
            report_html.build_html([], p, title="Empty")
            self.assertTrue(os.path.exists(p))
            # The donut degrades gracefully to a zero-state ring with no findings.
            with open(p, encoding="utf-8") as fh:
                self.assertIn("<svg", fh.read())




class ReportTest(unittest.TestCase):
    def test_workbook_builds_and_has_sheets(self):
        hosts = parser.parse_nmap_xml(SAMPLE)
        ad.analyze_hosts(hosts)
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "x.xlsx")
            build_workbook(hosts, out, meta={"subtitle": "t"})
            self.assertTrue(os.path.exists(out))
            sheets = xlsx.read_sheets(out)
        for name in ("Start Here", "Overview", "Checklist", "Services by Product",
                     "Vulnerabilities", "AD Quick Wins"):
            self.assertIn(name, sheets)

    def test_opens_in_openpyxl_if_available(self):
        # Optional: proves the stdlib-written file parses in a real xlsx engine.
        try:
            from openpyxl import load_workbook
        except ImportError:
            self.skipTest("openpyxl not installed")
        hosts = parser.parse_nmap_xml(SAMPLE)
        ad.analyze_hosts(hosts)
        with tempfile.TemporaryDirectory() as d:
            out = os.path.join(d, "x.xlsx")
            build_workbook(hosts, out)
            wb = load_workbook(out)
            self.assertIn("Checklist", wb.sheetnames)
