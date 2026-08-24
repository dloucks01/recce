"""High-fidelity integration tests: the whole workflow end-to-end, with a hard
focus on correctness of the spreadsheet - that the RIGHT fields land on the RIGHT
IP row, that per-IP tracking never bleeds across hosts, and that re-scans/updates
preserve everything.

These drive the real parser -> store -> workbook writer -> read-back -> report
paths (no nmap needed) against the bundled sample scan, whose four hosts each
have a distinct fingerprint:

    10.0.10.10  dc01.corp.local  Windows Server 2019  88,389,445,3389  ms17-010
    10.0.10.25  ws01.corp.local  Windows 10 21H2      135,445,3389     (no vulns)
    10.0.20.5   web01            Linux 5.4            22,80,443        4 vulns
    10.0.20.6   web02            Linux 5.4            22,80,21,3306    ftp-anon
"""

import contextlib
import io
import os
import shutil
import sys
import tempfile
import unittest

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from recce import ad, parser, xlsx
from recce import tracking as tr
from recce.models import Host, Port, Vuln
from recce.report_excel import (build_workbook, read_key_order,
                                 read_workbook_edits, read_workbook_tracking,
                                 update_workbook, STATUS_WIP, STATUS_TODO)
from recce.store import Store
from recce.targets import _subnet_of

SAMPLE = os.path.join(os.path.dirname(parser.__file__), "sample_scan.xml")

# Ground-truth facts, keyed by IP, for cross-checking the spreadsheet.
FACTS = {
    "10.0.10.10": {"host": "dc01.corp.local", "os": "Windows Server 2019",
                   "ports": [88, 389, 445, 3389], "nvulns": 1},
    "10.0.10.25": {"host": "ws01.corp.local", "os": "Windows 10",
                   "ports": [135, 445, 3389], "nvulns": 0},
    "10.0.20.5": {"host": "web01", "os": "Linux",
                  "ports": [22, 80, 443], "nvulns": 4},
    "10.0.20.6": {"host": "web02", "os": "Linux",
                  "ports": [21, 22, 80, 3306], "nvulns": 1},
}


def sample_hosts():
    hosts = parser.parse_nmap_xml(SAMPLE)
    for h in hosts:
        h.subnet = _subnet_of(h.ip)
        h.enumerated = True
    ad.analyze_hosts(hosts)
    return hosts


def header_index(rows, *must_have):
    """Row index of the real column-header row (the first row that holds every
    token in must_have). A legend/note line can precede the header, so we locate
    it instead of assuming row 0."""
    for i, r in enumerate(rows):
        if all(tok in r for tok in must_have):
            return i
    return 0


def rows_by_ip(sheets, title):
    """Return (header, {ip: [row-as-dict, ...]}) for a sheet with an IP column.

    Skips collapsible group-header band rows (they carry a label in the IP column
    but no Key), so callers only see real data rows keyed by a bare IP."""
    rows = sheets[title]
    hidx = header_index(rows, "IP")
    hdr = rows[hidx]
    ipc = hdr.index("IP")
    kidx = hdr.index("Key") if "Key" in hdr else None
    out: dict = {}
    for r in rows[hidx + 1:]:
        if kidx is not None and (len(r) <= kidx or not r[kidx]):
            continue                       # group-header band row - not data
        if len(r) > ipc and r[ipc]:
            out.setdefault(str(r[ipc]), []).append(dict(zip(hdr, r)))
    return hdr, out




# Re-export helpers so a caller could `from _workflow_helpers import ...` too.
from _workflow_helpers import SAMPLE, FACTS, sample_hosts, header_index, rows_by_ip, _LOOT_LINUX, _LOOT_WIN, _GNMAP  # noqa: F401





class NmapImportTest(unittest.TestCase):
    def test_split_product_version(self):
        from recce.parser import _split_product_version
        self.assertEqual(_split_product_version("OpenSSH 8.2p1 Ubuntu"),
                         ("OpenSSH", "8.2p1"))
        self.assertEqual(_split_product_version("Apache httpd 2.4.49"),
                         ("Apache httpd", "2.4.49"))
        self.assertEqual(_split_product_version("Microsoft Windows RPC"),
                         ("Microsoft Windows RPC", ""))

    def test_parse_gnmap(self):
        from recce import parser
        with tempfile.TemporaryDirectory() as d:
            f = os.path.join(d, "s.gnmap")
            with open(f, "w") as fh:
                fh.write(_GNMAP)
            hosts = parser.parse_gnmap(f)
            self.assertEqual(len(hosts), 1)
            h = hosts[0]
            self.assertEqual(h.ip, "10.0.20.6")
            self.assertIn("web02", h.hostnames)
            self.assertEqual({p.portid for p in h.ports}, {21, 22, 80})
            ftp = next(p for p in h.ports if p.portid == 21)
            self.assertEqual(ftp.product, "vsftpd")
            self.assertEqual(ftp.version, "2.3.4")
            self.assertEqual(h.os_family, "Linux")

    def test_parse_normal_text(self):
        from recce import parser
        normal = ("Nmap scan report for web02 (10.0.20.6)\n"
                  "Host is up (0.00042s latency).\n"
                  "PORT   STATE SERVICE VERSION\n"
                  "21/tcp open  ftp     vsftpd 2.3.4\n"
                  "80/tcp open  http    Apache httpd 2.4.49\n"
                  "445/tcp closed microsoft-ds\n")
        with tempfile.TemporaryDirectory() as d:
            f = os.path.join(d, "s.nmap")
            with open(f, "w") as fh:
                fh.write(normal)
            hosts = parser.parse_normal(f)
            self.assertEqual(len(hosts), 1)
            h = hosts[0]
            self.assertEqual(h.ip, "10.0.20.6")
            self.assertIn("web02", h.hostnames)
            self.assertEqual({p.portid for p in h.ports}, {21, 80})  # closed dropped
            ftp = next(p for p in h.ports if p.portid == 21)
            self.assertEqual((ftp.product, ftp.version), ("vsftpd", "2.3.4"))

    def test_parse_normal_bare_ip(self):
        from recce import parser
        with tempfile.TemporaryDirectory() as d:
            f = os.path.join(d, "s.nmap")
            with open(f, "w") as fh:
                fh.write("Nmap scan report for 10.0.0.9\n22/tcp open ssh\n")
            hosts = parser.parse_normal(f)
            self.assertEqual(hosts[0].ip, "10.0.0.9")
            self.assertEqual(hosts[0].hostnames, [])

    def test_parse_nmap_file_autodetects_all_formats(self):
        from recce import parser
        with tempfile.TemporaryDirectory() as d:
            # grepable content, no extension -> sniffed
            g = os.path.join(d, "noext_grep")
            with open(g, "w") as fh:
                fh.write(_GNMAP)
            self.assertTrue(parser.parse_nmap_file(g))
            # normal text, no extension -> sniffed
            n = os.path.join(d, "noext_normal")
            with open(n, "w") as fh:
                fh.write("Nmap scan report for 1.2.3.9\n80/tcp open http\n")
            self.assertEqual(parser.parse_nmap_file(n)[0].ip, "1.2.3.9")
            # xml
            x = os.path.join(d, "s.xml")
            with open(x, "w") as fh:
                fh.write('<?xml version="1.0"?><nmaprun><host><status state="up"/>'
                         '<address addr="1.2.3.4" addrtype="ipv4"/><ports>'
                         '<port protocol="tcp" portid="80"><state state="open"/>'
                         '<service name="http"/></port></ports></host></nmaprun>')
            self.assertEqual(parser.parse_nmap_file(x)[0].ip, "1.2.3.4")

    def test_masscan_xml_is_nmap_compatible(self):
        from recce import parser
        with tempfile.TemporaryDirectory() as d:
            f = os.path.join(d, "mass.xml")
            with open(f, "w") as fh:
                fh.write('<?xml version="1.0"?><nmaprun scanner="masscan"><host>'
                         '<address addr="10.0.30.9" addrtype="ipv4"/><ports>'
                         '<port protocol="tcp" portid="5432"><state state="open"/>'
                         '<service name="postgresql"/></port></ports></host></nmaprun>')
            hosts = parser.parse_nmap_file(f)
            self.assertEqual(hosts[0].ip, "10.0.30.9")
            self.assertEqual(hosts[0].open_ports[0].portid, 5432)

    def test_oa_directory_imports_once_prefers_xml(self):
        # A -oA set (base.xml + base.gnmap + base.nmap) must import once, from xml.
        from recce import cli
        with tempfile.TemporaryDirectory() as d:
            for ext, body in ((".gnmap", _GNMAP),
                              (".nmap", "Nmap scan report for x (10.0.20.6)\n"
                                        "21/tcp open ftp vsftpd 2.3.4\n"),
                              (".xml", '<?xml version="1.0"?><nmaprun><host>'
                                       '<status state="up"/><address addr="10.0.20.6" '
                                       'addrtype="ipv4"/><ports><port protocol="tcp" '
                                       'portid="21"><state state="open"/><service '
                                       'name="ftp" product="vsftpd" version="2.3.4"/>'
                                       '</port></ports></host></nmaprun>')):
                with open(os.path.join(d, "scan" + ext), "w") as fh:
                    fh.write(body)
            files = cli._collect_scan_files([d])
            self.assertEqual(len(files), 1)
            self.assertTrue(files[0].endswith(".xml"))

    def test_import_builds_workbook_with_checkmarks_and_findings(self):
        from recce import cli
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            f = os.path.join(d, "s.gnmap")
            with open(f, "w") as fh:
                fh.write(_GNMAP)
            eng = os.path.join(d, "eng")
            argv = ["import", f, "-o", eng, "--title", "T"]
            args = cli.build_arg_parser().parse_args(argv)
            with contextlib.redirect_stdout(io.StringIO()):
                self.assertEqual(args.func(args), 0)
            h = Store(os.path.join(eng, "results.sqlite")).get_host("10.0.20.6")
            self.assertTrue(h.enumerated)                  # checkmark set
            self.assertEqual(len(h.open_ports), 3)
            # offline version->CVE engine fired on the imported versions
            titles = " ".join(v.title for v in h.vulns)
            self.assertIn("vsftpd 2.3.4 backdoor", titles)
            self.assertIn("path traversal", titles.lower())
            self.assertTrue(os.path.exists(os.path.join(eng, "enumeration.xlsx")))

    def test_import_appends_subnets_and_merges_overlap(self):
        # Several scans (different subnets, then an overlapping host) must APPEND
        # new hosts and MERGE overlaps - never duplicate a host or a port.
        from recce import cli
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            eng = os.path.join(d, "eng")

            def imp(text):
                f = os.path.join(d, "s.gnmap")
                with open(f, "w") as fh:
                    fh.write(text)
                a = cli.build_arg_parser().parse_args(["import", f, "-o", eng])
                with contextlib.redirect_stdout(io.StringIO()):
                    a.func(a)

            imp("Host: 10.0.10.10 (dc01)\tPorts: 445/open/tcp//microsoft-ds//, "
                "3389/open/tcp//ms-wbt-server//\tIgnored State: closed\n"
                "Host: 10.0.10.25 (ws01)\tPorts: 445/open/tcp//microsoft-ds//"
                "\tIgnored State: closed\n")                       # subnet .10
            imp("Host: 10.0.20.5 (web01)\tPorts: 22/open/tcp//ssh//OpenSSH 8.2p1/"
                "\tIgnored State: closed\n")                       # subnet .20 (appended)
            imp("Host: 10.0.10.10 (dc01)\tPorts: 88/open/tcp//kerberos-sec//, "
                "445/open/tcp//microsoft-ds//\tIgnored State: closed\n")  # overlap
            s = Store(os.path.join(eng, "results.sqlite"))
            hosts = s.all_hosts()
            self.assertEqual(len(hosts), 3)                        # dc01 not duplicated
            self.assertEqual({h.subnet for h in hosts},
                             {"10.0.10.0/24", "10.0.20.0/24"})     # both subnets present
            dc = s.get_host("10.0.10.10")
            self.assertEqual(sorted(p.portid for p in dc.open_ports), [88, 445, 3389])
            s.close()

    def test_import_merges_and_preserves_tracking(self):
        from recce import cli
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            f = os.path.join(d, "s.gnmap")
            with open(f, "w") as fh:
                fh.write(_GNMAP)
            eng = os.path.join(d, "eng")
            def run():
                a = cli.build_arg_parser().parse_args(["import", f, "-o", eng])
                with contextlib.redirect_stdout(io.StringIO()):
                    a.func(a)
            run()
            # Tick via the `review` command (writes tracking AND regenerates the
            # workbook, the real path a tester's edits take).
            rv = cli.build_arg_parser().parse_args(
                ["review", "--host", "10.0.20.6", "--note", "manually reviewed",
                 "-o", eng])
            with contextlib.redirect_stdout(io.StringIO()):
                rv.func(rv)
            run()                                          # re-import same scan
            s = Store(os.path.join(eng, "results.sqlite"))
            h = s.get_host("10.0.20.6")
            self.assertEqual(len(h.open_ports), 3)         # not duplicated
            self.assertEqual(s.get_tracking().get("host:10.0.20.6"),
                             (True, "manually reviewed"))  # tick preserved
            s.close()




class IngestParserTest(unittest.TestCase):
    def test_parse_findings_and_skip_howto(self):
        from recce import ingest
        p = ingest.parse_loot(_LOOT_LINUX)
        self.assertTrue(p["is_recce"])
        self.assertEqual(p["hostname"], "web01")
        self.assertEqual(p["os"], "linux")
        texts = [f["text"] for f in p["findings"]]
        self.assertEqual(len(texts), 5)                       # 6 [!] lines, 1 in how-to
        self.assertTrue(all("MUST NOT BE INGESTED" not in t for t in texts))
        cats = {f["category"] for f in p["findings"]}
        self.assertTrue({"sudo", "suid", "kernel", "writable"} <= cats)

    def test_windows_detection(self):
        from recce import ingest
        p = ingest.parse_loot(_LOOT_WIN)
        self.assertEqual(p["os"], "windows")
        self.assertEqual(p["hostname"], "DBSRV01")

    def test_strips_ansi_colour(self):
        from recce import ingest
        coloured = ("recce-enum  host=h1\n\x1b[1;36m==== Sudo ====\x1b[0m\n"
                    "\x1b[1;33m[!] NOPASSWD sudo entries present\x1b[0m\n")
        p = ingest.parse_loot(coloured)
        self.assertEqual(len(p["findings"]), 1)
        self.assertEqual(p["findings"][0]["text"], "NOPASSWD sudo entries present")

    def test_dedup(self):
        from recce import ingest
        dupe = ("recce-enum host=h\n==== Sudo ====\n[!] same\n[!] same\n")
        self.assertEqual(len(ingest.parse_loot(dupe)["findings"]), 1)

    def test_empty_and_garbage_input_no_crash(self):
        from recce import ingest
        for blob in ("", "\n\n\n", "not recce output at all\nrandom text\n",
                     "\x00\x01\x02 binary-ish \xff\xfe", "=" * 5000,
                     "[!] finding with no banner and no section\n"):
            p = ingest.parse_loot(blob)          # must not raise
            self.assertIn("findings", p)
            self.assertIsInstance(p["findings"], list)

    def test_findings_without_banner_still_parse(self):
        from recce import ingest
        p = ingest.parse_loot("==== Sudo ====\n[!] NOPASSWD present\n")
        self.assertFalse(p["is_recce"])          # no banner
        self.assertEqual(len(p["findings"]), 1)  # but [!] lines still harvested

    def test_malformed_section_headers_tolerated(self):
        from recce import ingest
        # Ragged '=' fences and stray '=' in finding text must not break parsing.
        blob = ("recce-enum host=h\n=== Weird ==\n[!] a = b = c finding\n"
                "======\n[!] another\n")
        p = ingest.parse_loot(blob)
        self.assertEqual(len(p["findings"]), 2)




class IngestCommandTest(unittest.TestCase):
    def _eng(self, d, host=None):
        from recce.store import Store
        os.makedirs(os.path.join(d, "raw"), exist_ok=True)
        s = Store(os.path.join(d, "results.sqlite"))
        s.set_meta("engagement", "T")
        if host:
            s.upsert_host(host)
        s.close()

    def _ingest(self, d, loot_text, extra=None):
        from recce import cli
        loot = os.path.join(d, "loot.txt")
        with open(loot, "w") as fh:
            fh.write(loot_text)
        argv = ["ingest", loot, "-o", d] + (extra or [])
        args = cli.build_arg_parser().parse_args(argv)
        with contextlib.redirect_stdout(io.StringIO()):
            rc = args.func(args)
        return rc

    def test_ingest_folds_network_topology_and_survives_merge(self):
        from recce.store import Store
        loot = ("recce-enum host=web01 user=root now\n"
                "==== NETWORK ====\n"
                "NET-IFACE eth0 10.0.20.5/24\nNET-IFACE eth1 10.0.10.9/24\n"
                "NET-NEIGH 10.0.10.10 aa:bb:cc:00:00:10\n"
                "NET-PEER 10.0.10.10:445 ESTAB\n==== END NETWORK ====\n")
        with tempfile.TemporaryDirectory() as d:
            self._eng(d, Host(ip="10.0.20.5", hostnames=["web01"], enumerated=True))
            # topology-only loot (no [!] findings) must still ingest
            self.assertEqual(self._ingest(d, loot, ["--host", "10.0.20.5"]), 0)
            db = os.path.join(d, "results.sqlite")
            h = Store(db).get_host("10.0.20.5")
            self.assertEqual(len(h.topology["interfaces"]), 2)
            self.assertIn("10.0.10.10", h.topology["neighbors"])
            # a later, unrelated upsert (merge=True) must not drop the topology
            s = Store(db)
            again = s.get_host("10.0.20.5")
            again.vulns = []                          # simulate a re-scan touch
            s.upsert_host(again)                      # merge path
            s.close()
            h2 = Store(db).get_host("10.0.20.5")
            self.assertEqual(len(h2.topology.get("interfaces", [])), 2)

    def test_ingest_auto_resolves_host_from_own_interface_ip(self):
        # No --host: the box's own NET-IFACE IP must land the loot on the real
        # enumerated host in scope, not synthesize a local: entry.
        from recce.store import Store
        loot = ("recce-enum host=web01 user=root now\n"
                "==== NETWORK ====\n"
                "NET-IFACE eth0 10.0.20.5/24\n"
                "NET-NEIGH 10.0.10.10 aa:bb:cc:00:00:10\n"
                "==== END NETWORK ====\n")
        with tempfile.TemporaryDirectory() as d:
            self._eng(d, Host(ip="10.0.20.5", enumerated=True))   # in scope, no hostname
            self.assertEqual(self._ingest(d, loot), 0)            # no --host
            s = Store(os.path.join(d, "results.sqlite"))
            h = s.get_host("10.0.20.5")
            self.assertIsNotNone(h)
            self.assertEqual(len(h.topology["interfaces"]), 1)    # folded onto the real host
            self.assertIn("web01", h.hostnames)                   # banner hostname recorded
            self.assertIsNone(s.get_host("local:web01"))          # nothing synthesized

    def test_ingest_matches_host_by_hostname(self):
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            self._eng(d, Host(ip="10.0.0.50", hostnames=["web01"],
                              os_family="Linux", enumerated=True))
            self.assertEqual(self._ingest(d, _LOOT_LINUX), 0)
            h = Store(os.path.join(d, "results.sqlite")).get_host("10.0.0.50")
            self.assertEqual(len(h.local_findings), 5)
            self.assertTrue(h.privesc_checked)

    def test_ingest_is_idempotent(self):
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            self._eng(d, Host(ip="10.0.0.50", hostnames=["web01"]))
            self._ingest(d, _LOOT_LINUX)
            self._ingest(d, _LOOT_LINUX)               # re-ingest same loot
            h = Store(os.path.join(d, "results.sqlite")).get_host("10.0.0.50")
            self.assertEqual(len(h.local_findings), 5)  # not doubled

    def test_ingest_synthesizes_host_when_unknown(self):
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            self._eng(d)                                # empty engagement
            self.assertEqual(self._ingest(d, _LOOT_WIN), 0)
            h = Store(os.path.join(d, "results.sqlite")).get_host("local:DBSRV01")
            self.assertIsNotNone(h)
            self.assertEqual(h.os_family, "Windows")
            self.assertEqual(len(h.local_findings), 2)

    def test_ingest_host_flag_records_hostname(self):
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            self._eng(d, Host(ip="10.0.0.50"))          # no hostname stored
            self._ingest(d, _LOOT_LINUX, extra=["--host", "10.0.0.50"])
            h = Store(os.path.join(d, "results.sqlite")).get_host("10.0.0.50")
            # hostname from the loot banner is recorded, so a later no---host
            # ingest of the same box matches this entry instead of synthesizing.
            self.assertIn("web01", h.hostnames)

    def test_ingest_dedups_incoming_rows_on_new_host(self):
        # Two sections that map to the same category with identical finding text
        # must not create duplicate rows, even on a brand-new (unmerged) host.
        from recce.store import Store
        loot = ("recce-enum host=h1\n"
                "==== SUID / SGID / capabilities ====\n[!] same finding text\n"
                "==== Capabilities ====\n[!] same finding text\n")
        with tempfile.TemporaryDirectory() as d:
            self._eng(d)                                 # empty -> synthetic host
            self._ingest(d, loot, extra=["--host", "10.0.0.1"])
            h = Store(os.path.join(d, "results.sqlite")).get_host("10.0.0.1")
            self.assertEqual(len(h.local_findings), 1)

    def test_exploitation_sheet_lists_confirmed_findings(self):
        from recce.report_excel import build_workbook
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not installed (test-only dependency)")
        with tempfile.TemporaryDirectory() as d:
            self._eng(d, Host(ip="10.0.0.50", hostnames=["web01"], os_family="Linux"))
            self._ingest(d, _LOOT_LINUX)      # sudo/suid/shadow -> confirmed
            from recce.store import Store
            hosts = Store(os.path.join(d, "results.sqlite")).all_hosts()
            p = os.path.join(d, "wb.xlsx")
            build_workbook(hosts, p)
            wb = openpyxl.load_workbook(p)
            self.assertIn("Exploitation", wb.sheetnames)
            ws = wb["Exploitation"]
            hdr = [c.value for c in ws[1]]
            ti = hdr.index("Existing tool")
            tools = " ".join(str(r[ti]) for r in ws.iter_rows(min_row=2, values_only=True)
                             if r[ti])
            self.assertIn("GTFOBins", tools)      # sudo / SUID findings
            self.assertGreaterEqual(ws.max_row - 1, 2)

    def test_high_signal_findings_promoted_to_vulns(self):
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            self._eng(d, Host(ip="10.0.0.50", hostnames=["web01"], os_family="Linux"))
            self._ingest(d, _LOOT_LINUX)
            h = Store(os.path.join(d, "results.sqlite")).get_host("10.0.0.50")
            local_vulns = [v for v in h.vulns if v.source == "local"]
            titles = " ".join(v.title for v in local_vulns)
            self.assertTrue(local_vulns)                    # some got promoted
            self.assertIn("Sudo misconfiguration", titles)  # NOPASSWD / ALL
            self.assertIn("Readable /etc/shadow", titles)
            # Promoted vulns are confirmed local observations with a CWE.
            self.assertTrue(all(v.confidence == "confirmed" and v.cwes
                                for v in local_vulns))

    def test_promotion_is_idempotent(self):
        from recce.store import Store
        with tempfile.TemporaryDirectory() as d:
            self._eng(d, Host(ip="10.0.0.50", hostnames=["web01"], os_family="Linux"))
            self._ingest(d, _LOOT_LINUX)
            self._ingest(d, _LOOT_LINUX)          # re-ingest
            h = Store(os.path.join(d, "results.sqlite")).get_host("10.0.0.50")
            local = [v for v in h.vulns if v.source == "local"]
            self.assertEqual(len(local), len({v.title for v in local}))  # no dupes

    def test_ingested_findings_appear_on_privesc_sheet(self):
        with tempfile.TemporaryDirectory() as d:
            self._eng(d, Host(ip="10.0.0.50", hostnames=["web01"], os_family="Linux"))
            self._ingest(d, _LOOT_LINUX)
            try:
                import openpyxl
            except ImportError:
                self.skipTest("openpyxl not installed (test-only dependency)")
            ws = openpyxl.load_workbook(os.path.join(d, "enumeration.xlsx"))["Priv-Esc"]
            hdr = [c.value for c in ws[1]]
            ti = hdr.index("Type")
            rows = list(ws.iter_rows(min_row=2, values_only=True))
            # The 5 ingested findings each become a row verdicted as an escalation
            # path or an observation (this fresh host has no remote findings).
            on_target = sum(1 for r in rows if r[ti] in ("Escalation path", "Finding"))
            self.assertEqual(on_target, 5)
            # ...and at least some are verdicted as actual escalation paths.
            self.assertGreater(sum(1 for r in rows if r[ti] == "Escalation path"), 0)
            # The Priv-Esc tab is now findings-only: NO generic checklist rows
            # (a swept host also gets no 'run recce deploy' to-do).
            self.assertEqual(sum(1 for r in rows if r[ti] in ("Checklist", "To do")), 0)
            # The generic OS checklist lives on the separate reference sheet.
            wb = openpyxl.load_workbook(os.path.join(d, "enumeration.xlsx"))
            self.assertIn("Priv-Esc Playbook", wb.sheetnames)
            pb = wb["Priv-Esc Playbook"]
            self.assertGreater(pb.max_row, 1)         # has playbook rows

    def test_triaged_vuln_counts_toward_coverage(self):
        """Regression: the Vulnerabilities sheet's row key and the coverage
        counter's key must be identical, or ticking Triaged is never counted."""
        from recce import tracking as tr
        from recce.models import Vuln
        from recce.report_excel import _spec_vulns
        h = Host(ip="10.0.0.5", ports=[Port(portid=445, service="microsoft-ds")],
                 vulns=[Vuln(ip="10.0.0.5", port=445, protocol="tcp",
                             script_id="smb-vuln-ms17-010", title="ms17-010 RCE",
                             severity="high", source="nse")])
        sheet_key = _spec_vulns([h]).rows[0]["key"]
        # the sheet key, the canonical key, and the coverage key all agree
        self.assertEqual(sheet_key, tr.vuln_row_key(h.vulns[0]))
        self.assertIn(sheet_key, tr.item_keys([h])["vulns"])
        # untriaged -> 0/1; triaging the sheet's key -> 1/1 (was stuck at 0 before)
        self.assertEqual(tr.compute_coverage([h], {})["vulns"],
                         {"total": 1, "done": 0, "pct": 0})
        cov = tr.compute_coverage([h], {sheet_key: (True, "")})["vulns"]
        self.assertEqual(cov, {"total": 1, "done": 1, "pct": 100})

    def test_vuln_row_key_matches_store_dedup_granularity(self):
        """Regression: the workbook/coverage key must not truncate the title
        more coarsely than the store's dedup key (models.Vuln.key uses [:60]),
        or two store-distinct findings collapse to one Vulnerabilities row and
        coverage undercounts."""
        from recce import tracking as tr
        from recce.models import Vuln
        from recce.report_excel import _spec_vulns
        # Two findings identical for 40 chars, differing only at chars 41-60.
        base = "Apache httpd 2.4.49 Path Traversal RCE - "  # 41 chars
        v1 = Vuln(ip="10.0.0.5", port=443, protocol="tcp", script_id="version-db",
                  title=base + "CVE-2021-41773", severity="high", source="db")
        v2 = Vuln(ip="10.0.0.5", port=443, protocol="tcp", script_id="version-db",
                  title=base + "CVE-2021-42013", severity="high", source="db")
        # Store keeps both distinct (its key uses title[:60])...
        self.assertNotEqual(v1.key, v2.key)
        # ...so the workbook keys must also be distinct (no collapse).
        self.assertNotEqual(tr.vuln_row_key(v1), tr.vuln_row_key(v2))
        h = Host(ip="10.0.0.5", ports=[Port(portid=443, service="https")],
                 vulns=[v1, v2])
        rows = _spec_vulns([h]).rows
        self.assertEqual(len(rows), 2, "both findings must appear on the sheet")
        self.assertEqual(tr.compute_coverage([h], {})["vulns"]["total"], 2)
