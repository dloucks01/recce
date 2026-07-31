"""Scale test: a large multi-subnet datastore, end to end through the CLI.

Multi-subnet is recce's headline use case, but the rest of the suite never exercises
more than a handful of hosts. This builds a realistic large engagement (hundreds of
hosts across many subnets, each with ports + findings), then rebuilds the workbook
and runs `status` and `access`, asserting the outputs are correct and the whole thing
stays well clear of an O(n^2) time blow-up.
"""
import contextlib
import io
import os
import shutil
import time
import unittest

from recce import cli
from recce.models import Host, Port, Vuln
from recce.store import Store

_SUBNETS = 20
_PER = 25                      # 500 hosts total
_TOTAL = _SUBNETS * _PER
_TIME_BUDGET = 90.0            # generous; a quadratic build would blow past this

_SVCS = [(22, "ssh", "OpenSSH 8.2p1"), (80, "http", "Apache httpd 2.4.41"),
         (445, "microsoft-ds", "Windows Server 2019"), (3389, "ms-wbt-server", "")]


def _build(path):
    store = Store(path)
    store.set_meta("engagement", "Scale Test")
    for s in range(_SUBNETS):
        subnet = f"10.{s}.0.0/24"
        for h in range(_PER):
            ip = f"10.{s}.0.{h + 1}"
            ports = [Port(portid=p, protocol="tcp", state="open", service=sv,
                          product=pr, vuln_scanned=True) for p, sv, pr in _SVCS]
            vulns = [Vuln(ip=ip, port=445, protocol="tcp",
                          script_id="smb-vuln-ms17-010", state="VULNERABLE",
                          title="MS17-010 (EternalBlue)", severity="critical",
                          source="nse", cwes=["CWE-119"])]
            host = Host(ip=ip, subnet=subnet, state="up", up_reason="syn-ack",
                        enumerated=True, hostnames=[f"host{s}-{h}"],
                        os_name="Windows" if h % 2 else "Linux",
                        ports=ports, vulns=vulns)
            store.upsert_host(host, merge=False)
    return store


class ScaleTest(unittest.TestCase):

    def setUp(self):
        self.dir = os.path.join(os.environ.get("TMPDIR", "/tmp"),
                                f"recce_scale_{os.getpid()}")
        shutil.rmtree(self.dir, ignore_errors=True)
        os.makedirs(self.dir)

    def tearDown(self):
        shutil.rmtree(self.dir, ignore_errors=True)

    def test_large_engagement_reports_status_and_access(self):
        db = os.path.join(self.dir, "results.sqlite")

        t0 = time.time()
        store = _build(db)
        store.close()
        build_s = time.time() - t0
        self.assertEqual(len(Store(db).all_hosts()), _TOTAL)

        # Rebuild all reports from the big datastore.
        t0 = time.time()
        with contextlib.redirect_stdout(io.StringIO()):
            rc = cli.main(["report", "-o", self.dir])
        gen_s = time.time() - t0
        self.assertEqual(rc, 0)

        # The workbook is valid and carries a row per host (plus per-subnet bands).
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl not installed (test-only dependency)")
        wb = openpyxl.load_workbook(os.path.join(self.dir, "enumeration.xlsx"))
        self.assertIn("Checklist", wb.sheetnames)
        self.assertGreaterEqual(wb["Checklist"].max_row, _TOTAL)
        for name in ("enumeration.md", "report.html", "services.csv"):
            self.assertGreater(
                os.path.getsize(os.path.join(self.dir, name)), 0)

        # status and access both run cleanly over the whole set.
        with contextlib.redirect_stdout(io.StringIO()) as out:
            self.assertEqual(cli.main(["status", "-o", self.dir]), 0)
        self.assertIn("OVERALL", out.getvalue())
        with contextlib.redirect_stdout(io.StringIO()):
            self.assertEqual(cli.main(["access", "-o", self.dir]), 0)

        # Guardrail: the build + report of 500 hosts must stay near-linear.
        self.assertLess(build_s + gen_s, _TIME_BUDGET,
                        f"scale too slow: build={build_s:.1f}s gen={gen_s:.1f}s "
                        f"for {_TOTAL} hosts (possible O(n^2))")


if __name__ == "__main__":
    unittest.main()
